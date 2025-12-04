"""
잔여객실 현황 독립 대시보드
- 개발자 없이 실행 가능
- DB 직접 연결
- 웹 브라우저에서 실행
"""

import streamlit as st
import streamlit.components.v1 as components
import pyodbc
import pandas as pd
import json
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go

# set_page_config는 반드시 첫 번째 Streamlit 명령이어야 함
st.set_page_config(page_title="여객 현황 대시보드", layout="wide", initial_sidebar_state="collapsed")

# DB 설정
try:
    DB_CONFIG = {
        'server': st.secrets["database"]["server"],
        'base_database': st.secrets["database"]["base_database"],
        'cruise_database': st.secrets["database"]["cruise_database"],
        'username': st.secrets["database"]["username"],
        'password': st.secrets["database"]["password"],
    }
except Exception as e:
    st.error("DB 설정을 찾을 수 없습니다. .streamlit/secrets.toml 파일을 확인하세요.")
    DB_CONFIG = {
        'server': '',
        'base_database': '',
        'cruise_database': '',
        'username': '',
        'password': '',
    }

# ============================================================
# JavaScript 기반 모달 (페이지 새로고침 없음)
# ============================================================

# Session state 초기화 (필요시)
st.markdown("""
<div style="text-align: center; padding: 40px 0 30px 0;">
    <h1 style="font-size: 32px; font-weight: 700; letter-spacing: -0.5px; color: #232A5E; margin-bottom: 8px; font-family: 'Noto Sans KR', sans-serif;">
        NEOHELIOS CRUISE
    </h1>
    <p style="font-size: 14px; font-weight: 400; color: #88949C; letter-spacing: -0.5px; font-family: 'Noto Sans KR', sans-serif;">
        여객 현황 대시보드
    </p>
</div>
""", unsafe_allow_html=True)

# NEOHELIOS 디자인 시스템 스타일
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');
    
    /* === CSS Variables (Design Tokens) === */
    :root {
        --nh-primary: #436CFC;
        --nh-primary-light: #F3F6FF;
        --nh-helios-blue: #232A5E;
        --nh-alert-red: #EA3336;
        --nh-bg-gray: #F3F7F9;
        --nh-bg-white: #FFFFFF;
        --nh-border: #DAE0E3;
        --nh-text-dark: #0E0E2C;
        --nh-text-gray: #88949C;
        --nh-text-light: #9EA8B0;
        --nh-text-disabled: #B7BFC5;
        --nh-disabled-bg: #E3E8EB;
        --nh-font: 'Noto Sans KR', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    
    /* === 전체 배경 === */
    .stApp {
        background: var(--nh-bg-gray);
        color: var(--nh-text-dark);
        font-family: var(--nh-font);
    }
    
    /* === 메인 컨테이너 === */
    .main {
        background: var(--nh-bg-gray);
        padding: 1rem 2rem;
        max-width: 100%;
        margin: 0;
    }
    
    .block-container {
        padding-left: 1rem !important;
        padding-right: 1rem !important;
        padding-top: 1rem !important;
        padding-bottom: 0rem !important;
        max-width: 100% !important;
    }
    
    section.main > div {
        max-width: 100%;
        padding-left: 2rem;
        padding-right: 2rem;
    }
    
    /* === 제목 스타일 === */
    h1 {
        color: var(--nh-helios-blue);
        font-family: var(--nh-font);
        font-weight: 700;
        letter-spacing: -0.5px;
        margin-bottom: 10px;
    }
    
    h3 {
        color: var(--nh-helios-blue);
        font-family: var(--nh-font);
        font-weight: 500;
        font-size: 14px;
        letter-spacing: -0.5px;
        margin-bottom: 16px;
    }
    
    /* === 버튼 스타일 - Primary (SEABLUE) === */
    .stButton > button {
        background: var(--nh-primary);
        color: var(--nh-bg-white);
        font-family: var(--nh-font);
        font-weight: 500;
        border: none;
        border-radius: 5px;
        padding: 12px 24px;
        font-size: 14px;
        letter-spacing: -0.5px;
        transition: all 0.2s ease;
        box-shadow: none;
    }
    
    .stButton > button:hover {
        background: #3459E6;
        box-shadow: 0 2px 8px rgba(67, 108, 252, 0.3);
        transform: translateY(-1px);
    }
    
    /* === 다운로드 버튼 - Secondary === */
    .stDownloadButton > button {
        background: var(--nh-bg-white);
        color: var(--nh-helios-blue);
        font-family: var(--nh-font);
        border: 1px solid var(--nh-border);
        font-weight: 500;
        border-radius: 5px;
        padding: 10px 20px;
        font-size: 14px;
        letter-spacing: -0.5px;
        width: auto;
        transition: all 0.2s ease;
    }
    
    .stDownloadButton > button:hover {
        background: var(--nh-primary);
        color: var(--nh-bg-white);
        border-color: var(--nh-primary);
    }
    
    /* === 입력 필드 스타일 === */
    .stSelectbox > div > div,
    .stDateInput > div > div,
    .stTextInput > div > div {
        border-radius: 5px;
        border: 1px solid var(--nh-border);
        background: var(--nh-bg-gray);
        transition: all 0.2s ease;
    }
    
    .stSelectbox > div > div:hover,
    .stDateInput > div > div:hover,
    .stTextInput > div > div:hover {
        border-color: var(--nh-text-light);
    }
    
    .stSelectbox > div > div:focus-within,
    .stDateInput > div > div:focus-within,
    .stTextInput > div > div:focus-within {
        border: 2px solid var(--nh-primary) !important;
        background: var(--nh-bg-gray);
    }
    
    /* === 라벨 스타일 === */
    .stSelectbox label,
    .stDateInput label,
    .stTextInput label {
        font-family: var(--nh-font);
        font-size: 12px;
        font-weight: 500;
        letter-spacing: -0.5px;
        color: var(--nh-helios-blue);
    }
    
    /* === 입력 필드 내 텍스트 === */
    .stSelectbox input,
    .stSelectbox select,
    .stSelectbox div[data-baseweb="select"] > div,
    .stSelectbox div[data-baseweb="select"] span,
    .stDateInput input,
    .stTextInput input {
        color: var(--nh-text-dark) !important;
        font-family: var(--nh-font) !important;
        font-size: 14px !important;
        background-color: var(--nh-bg-gray) !important;
        letter-spacing: -0.5px;
    }
    
    /* === Disabled 필드 === */
    .stTextInput input:disabled {
        color: var(--nh-text-disabled) !important;
        background-color: var(--nh-disabled-bg) !important;
        -webkit-text-fill-color: var(--nh-text-disabled) !important;
        opacity: 1 !important;
    }
    
    /* === 드롭다운 메뉴 === */
    [role="listbox"],
    [data-baseweb="menu"],
    [data-baseweb="popover"] > div,
    .stSelectbox [role="option"] {
        background-color: var(--nh-bg-white) !important;
        border: 1px solid var(--nh-border);
        border-radius: 5px;
    }
    
    /* === 드롭다운 옵션 === */
    [role="option"],
    [data-baseweb="menu"] li,
    .stSelectbox li {
        background-color: var(--nh-bg-white) !important;
        color: var(--nh-text-dark) !important;
        font-family: var(--nh-font) !important;
        font-size: 14px !important;
        letter-spacing: -0.5px;
    }
    
    [role="option"]:hover,
    [data-baseweb="menu"] li:hover {
        background-color: var(--nh-primary-light) !important;
        color: var(--nh-text-dark) !important;
    }
    
    /* === Success/Info 메시지 === */
    .stSuccess {
        background: #E8F5E9;
        border-left: 4px solid #4CAF50;
        padding: 12px 16px;
        border-radius: 5px;
        font-family: var(--nh-font);
    }
    
    .stInfo {
        background: var(--nh-primary-light);
        border-left: 4px solid var(--nh-primary);
        padding: 12px 16px;
        border-radius: 5px;
        font-family: var(--nh-font);
    }
    
    /* === 탭 스타일 === */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0;
        background-color: transparent;
        border-bottom: 1px solid var(--nh-border);
        padding-bottom: 0;
        margin-bottom: 24px;
    }
    
    .stTabs [data-baseweb="tab"] {
        background-color: transparent !important;
        border: none !important;
        border-bottom: 3px solid transparent !important;
        border-radius: 0 !important;
        padding: 16px 24px !important;
        transition: all 0.2s ease;
    }
    
    .stTabs button[data-baseweb="tab"] {
        color: var(--nh-text-gray) !important;
        font-family: var(--nh-font) !important;
        font-size: 16px !important;
        font-weight: 500 !important;
        letter-spacing: -0.5px;
    }
    
    .stTabs [data-baseweb="tab"]:hover button {
        color: var(--nh-helios-blue) !important;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: transparent !important;
        border-bottom: 3px solid var(--nh-primary) !important;
    }
    
    .stTabs [aria-selected="true"] button {
        color: var(--nh-primary) !important;
        font-weight: 700 !important;
    }
    
    .stTabs [data-baseweb="tab-panel"] {
        padding-top: 24px;
    }
    
    /* === 반응형 테이블 컨테이너 === */
    .responsive-table-container {
        overflow-x: auto;
        -webkit-overflow-scrolling: touch;
        margin: 16px 0;
        width: 100%;
        max-width: 100%;
        border-radius: 5px;
        border: 1px solid var(--nh-border);
        background: var(--nh-bg-white);
    }
    
    .responsive-table-container table {
        min-width: 100%;
        width: 100%;
    }
    
    /* === 클릭 가능한 셀 === */
    .clickable-cell {
        cursor: pointer;
        transition: all 0.2s ease;
    }
    
    .clickable-cell:hover {
        opacity: 0.8;
        transform: scale(1.02);
    }
    
    /* === 모달 스타일 === */
    .modal {
        display: none;
        position: fixed;
        z-index: 9999;
        left: 0;
        top: 0;
        width: 100%;
        height: 100%;
        overflow: auto;
        background-color: rgba(14, 14, 44, 0.5);
        animation: fadeIn 0.2s ease;
    }
    
    .modal-content {
        background-color: var(--nh-bg-white);
        margin: 5% auto;
        padding: 0;
        border-radius: 5px;
        width: 90%;
        max-width: 900px;
        box-shadow: 0 4px 24px rgba(14, 14, 44, 0.15);
        animation: slideDown 0.2s ease;
    }
    
    .modal-header {
        background: var(--nh-helios-blue);
        color: white;
        padding: 16px 20px;
        border-radius: 5px 5px 0 0;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    
    .modal-header h2 {
        margin: 0;
        font-family: var(--nh-font);
        font-size: 16px;
        font-weight: 700;
        letter-spacing: -0.5px;
    }
    
    .modal-body {
        padding: 20px;
        max-height: 60vh;
        overflow-y: auto;
    }
    
    .close-modal {
        color: white;
        font-size: 24px;
        font-weight: bold;
        cursor: pointer;
        background: none;
        border: none;
        padding: 0;
        width: 28px;
        height: 28px;
        line-height: 28px;
        text-align: center;
        transition: all 0.2s ease;
        border-radius: 4px;
    }
    
    .close-modal:hover {
        background: rgba(255, 255, 255, 0.2);
    }
    
    .modal-info {
        background: var(--nh-bg-gray);
        padding: 16px;
        border-radius: 5px;
        margin-bottom: 16px;
        border-left: 4px solid var(--nh-helios-blue);
    }
    
    .modal-info-item {
        display: inline-block;
        margin-right: 24px;
        font-size: 14px;
        font-family: var(--nh-font);
    }
    
    .modal-info-label {
        font-weight: 500;
        color: var(--nh-helios-blue);
    }
    
    .modal-info-value {
        color: var(--nh-text-gray);
        margin-left: 8px;
    }
    
    .modal-table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 12px;
    }
    
    .modal-table th {
        background: var(--nh-helios-blue);
        color: #FAFCFE;
        padding: 10px 12px;
        text-align: left;
        font-family: var(--nh-font);
        font-weight: 700;
        font-size: 12px;
        letter-spacing: -0.5px;
    }
    
    .modal-table td {
        padding: 10px 12px;
        border-bottom: 1px solid var(--nh-border);
        color: var(--nh-text-dark);
        font-family: var(--nh-font);
        font-size: 14px;
        letter-spacing: -0.5px;
    }
    
    .modal-table tr:hover {
        background: var(--nh-bg-gray);
    }
    
    .loading-spinner {
        text-align: center;
        padding: 40px;
        color: var(--nh-text-gray);
        font-family: var(--nh-font);
    }
    
    @keyframes fadeIn {
        from { opacity: 0; }
        to { opacity: 1; }
    }
    
    @keyframes slideDown {
        from { 
            transform: translateY(-20px);
            opacity: 0;
        }
        to { 
            transform: translateY(0);
            opacity: 1;
        }
    }
    
    /* === 모바일 최적화 (768px 이하) === */
    @media screen and (max-width: 768px) {
        h1 {
            font-size: 20px !important;
        }
        
        h3 {
            font-size: 12px !important;
        }
        
        .main {
            padding: 0.5rem;
        }
        
        .responsive-table-container table {
            font-size: 12px;
        }
        
        .responsive-table-container th,
        .responsive-table-container td {
            padding: 8px 6px !important;
            font-size: 12px !important;
        }
        
        .responsive-table-container th[rowspan] {
            font-size: 12px !important;
            padding: 8px 6px !important;
        }
        
        .legend-container {
            grid-template-columns: 1fr !important;
            gap: 8px !important;
        }
        
        .legend-container > div {
            font-size: 12px !important;
        }
        
        .legend-container span[style*="width: 24px"] {
            width: 18px !important;
            height: 18px !important;
        }
        
        .modal-content {
            width: 95%;
            margin: 10% auto;
        }
        
        .modal-header h2 {
            font-size: 16px;
        }
        
        .modal-body {
            padding: 16px;
            max-height: 70vh;
        }
        
        .modal-info {
            padding: 12px;
        }
        
        .modal-info-item {
            display: block;
            margin-right: 0;
            margin-bottom: 8px;
            font-size: 13px;
        }
        
        .modal-table th,
        .modal-table td {
            padding: 8px 6px !important;
            font-size: 12px !important;
        }
    }
    
    /* 작은 모바일 (480px 이하) */
    @media screen and (max-width: 480px) {
        h1 {
            font-size: 20px !important;
        }
        
        .responsive-table-container table {
            font-size: 10px;
        }
        
        .responsive-table-container th,
        .responsive-table-container td {
            padding: 6px 3px !important;
            font-size: 10px !important;
        }
        
        .responsive-table-container th[rowspan] {
            font-size: 11px !important;
            padding: 8px 4px !important;
        }
        
        /* 버튼 크기 조정 */
        .stButton > button,
        .stDownloadButton > button {
            padding: 12px 16px;
            font-size: 13px;
        }
        
        /* 모달 작은 모바일 최적화 */
        .modal-content {
            width: 98%;
            margin: 5% auto;
        }
        
        .modal-header {
            padding: 16px;
        }
        
        .modal-header h2 {
            font-size: 14px;
        }
        
        .close-modal {
            font-size: 28px;
        }
        
        .modal-body {
            padding: 12px;
        }
        
        .modal-info-item {
            font-size: 12px;
        }
        
        .modal-table th,
        .modal-table td {
            padding: 6px 4px !important;
            font-size: 11px !important;
        }
    }
</style>
""", unsafe_allow_html=True)

# DB 연결 (필터용 데이터 조회)
try:
    drivers_to_try = [
        '{ODBC Driver 17 for SQL Server}',
        '{ODBC Driver 18 for SQL Server}',
        '{SQL Server}',
        '{SQL Server Native Client 11.0}'
    ]
    
    driver = None
    for d in drivers_to_try:
        try:
            conn_string_base = f"DRIVER={d};SERVER={DB_CONFIG['server']};DATABASE={DB_CONFIG['base_database']};UID={DB_CONFIG['username']};PWD={DB_CONFIG['password']}"
            test_conn = pyodbc.connect(conn_string_base)
            test_conn.close()
            driver = d
            break
        except:
            continue
    
    if driver:
        # Vessels 조회
        conn_base = pyodbc.connect(conn_string_base)
        vessels_query = "SELECT id, code, name FROM vessels WHERE deleted_at IS NULL AND is_cruise_available = 1 ORDER BY name"
        df_vessels = pd.read_sql(vessels_query, conn_base)
        
        # Routes 조회
        routes_query = "SELECT id, code, description FROM routes WHERE deleted_at IS NULL ORDER BY code"
        df_routes = pd.read_sql(routes_query, conn_base)
        
        # Ports 조회 (컬럼명 확인 필요 - 일단 주석 처리)
        # ports_query = "SELECT id, name FROM ports WHERE deleted_at IS NULL ORDER BY name"
        # df_ports = pd.read_sql(ports_query, conn_base)
        df_ports = pd.DataFrame()  # 임시로 비활성화
        conn_base.close()
    else:
        st.error("❌ ODBC 드라이버를 찾을 수 없습니다.")
        df_vessels = pd.DataFrame()
        df_routes = pd.DataFrame()
        df_ports = pd.DataFrame()
except Exception as e:
    st.warning(f"⚠️ 필터 데이터 로딩 실패: {str(e)}")
    df_vessels = pd.DataFrame()
    df_routes = pd.DataFrame()
    df_ports = pd.DataFrame()

# 필터 섹션
st.markdown('<h3 style="color: #2d2d2d; font-weight: 600; font-size: 16px; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 20px;">검색 조건</h3>', unsafe_allow_html=True)
col1, col2, col3, col4 = st.columns(4)

# 선박별 항로 매핑
vessel_routes = {
    'PSMC': ['BOC', 'ONC', 'KSC'],           # 크루즈선
    'PSTL': ['TSL'],                          # 고속선 (대마도)
    'PSGR': ['EAS', 'SCC', 'FWC', 'SND', 'NFW']  # 여객선
}

# 항로별 포트 매핑
route_ports = {
    # PSMC 항로
    'BOC': ['전체', 'PUS', 'OSA'],           # 부산-오사카
    'ONC': ['전체', 'PUS'],                   # 부산 주말 크루즈 (왕복)
    'KSC': ['전체', 'PUS'],                   # 한국해협 크루즈 (왕복)
    # PSTL 항로
    'TSL': ['전체', 'PUS', 'IZH', 'HTK'],    # 대마도 (부산-이즈하라-히타카츠)
    # PSGR 항로 (모두 부산 출도착)
    'EAS': ['전체', 'PUS'],                   # 동해
    'SCC': ['전체', 'PUS'],                   # 속초
    'FWC': ['전체', 'PUS'],                   # 불꽃크루즈
    'SND': ['전체', 'PUS'],                   # 선상디너
    'NFW': ['전체', 'PUS']                    # 야간불꽃
}

# TSL port_id 매핑 (proforma_schedules.port_id)
TSL_PORT_IDS = {
    'PUS': 1777,   # KRPUS - Busan
    'IZH': 1633,   # JPIZH - Izuhara
    'HTK': 3271    # JPHTK - Hitakatsu
}

# port_id → 포트 코드 역방향 매핑
PORT_CODE_MAP = {v: k for k, v in TSL_PORT_IDS.items()}
# PSMC용 추가 포트
PORT_CODE_MAP.update({
    1777: 'PUS',    # 부산
    1633: 'IZH',    # 이즈하라
    3271: 'HTK',    # 히타카츠
    1693: 'OSA',    # 오사카 (JPOSA)
    1746: 'FUK',    # 후쿠오카
})

# 출발지/도착지에 따른 direction 결정용 매핑
route_direction_map = {
    'BOC': {'first': 'PUS', 'second': 'OSA'},
    'ONC': {'first': 'PUS', 'second': 'PUS'},
    'KSC': {'first': 'PUS', 'second': 'PUS'},
    'TSL': {'first': 'PUS', 'second': 'IZH'},
    'EAS': {'first': 'PUS', 'second': 'FUK'},
    'SCC': {'first': 'PUS', 'second': 'FUK'},
    'FWC': {'first': 'PUS', 'second': 'FUK'},
    'SND': {'first': 'PUS', 'second': 'FUK'},
    'NFW': {'first': 'PUS', 'second': 'FUK'},
}

# 좌석 기반 선박 (1객실 = 1승객)
seat_based_vessels = ['PSTL', 'PSGR']

with col1:
    # 선박 선택 (첫 번째!)
    vessel_options = list(vessel_routes.keys())
    selected_vessel = st.selectbox("선박", vessel_options, index=0, key="vessel_select")

with col2:
    # 항로: 선박에 따라 동적 변경
    route_options = vessel_routes.get(selected_vessel, ['BOC'])
    selected_route = st.selectbox("항로", route_options, index=0, key="route_select")

with col3:
    # 출발지: 항로에 따라 동적 변경
    origin_options = route_ports.get(selected_route, ['전체', 'PUS', 'OSA'])
    selected_origin = st.selectbox("출발지", origin_options, index=0, key="origin_select")

with col4:
    # 도착지: 항로에 따라 동적 변경
    destination_options = route_ports.get(selected_route, ['전체', 'PUS', 'OSA'])
    selected_destination = st.selectbox("도착지", destination_options, index=0, key="destination_select")

col5, col6, col7 = st.columns([2, 2, 1])

with col5:
    start_date = st.date_input("출항시작일", datetime.today())

with col6:
    end_date = st.date_input("출항종료일", datetime.today() + timedelta(days=30))

with col7:
    st.markdown('<div style="margin-top: 28px;"></div>', unsafe_allow_html=True)
    query_button = st.button("조회", type="primary")

st.markdown('<hr style="border: none; height: 1px; background: #e0e0e0; margin: 30px 0;">', unsafe_allow_html=True)

# 조회 버튼 처리
if query_button:
    
    # DB 연결
    # 여러 드라이버 버전 자동 시도
    drivers_to_try = [
        "ODBC Driver 18 for SQL Server",
        "ODBC Driver 17 for SQL Server",
        "ODBC Driver 13 for SQL Server",
        "SQL Server Native Client 11.0",
        "SQL Server",
    ]
    
    driver = None
    for d in drivers_to_try:
        if d in pyodbc.drivers():
            driver = d
            break
    
    if not driver:
        st.error("❌ SQL Server ODBC 드라이버를 찾을 수 없습니다.")
        st.info("드라이버 설치 필요: https://go.microsoft.com/fwlink/?linkid=2249004")
        st.stop()
    
    conn_string_base = f"""
        Driver={{{driver}}};
        Server={DB_CONFIG['server']};
        Database={DB_CONFIG['base_database']};
        UID={DB_CONFIG['username']};
        PWD={DB_CONFIG['password']};
    """
    
    conn_string_cruise = f"""
        Driver={{{driver}}};
        Server={DB_CONFIG['server']};
        Database={DB_CONFIG['cruise_database']};
        UID={DB_CONFIG['username']};
        PWD={DB_CONFIG['password']};
    """
    
    with st.spinner('데이터 조회 중...'):
        try:
            # 1. 스케줄 조회 (neohelios_base)
            conn_base = pyodbc.connect(conn_string_base)
            
            # 항로 코드를 route_id로 변환
            route_map = {
                'BOC': 1, 'ONC': 2, 'KSC': 3, 'TSL': 5,
                'EAS': 7, 'SCC': 8, 'FWC': 9, 'SND': 10, 'NFW': 11
            }
            selected_route_id = route_map.get(selected_route, 1)
            
            # TSL은 특별 처리 (arrival_schedule_id의 port로 필터링)
            is_tsl = (selected_route == 'TSL')
            
            # 출발지/도착지에 따른 direction 결정 (TSL 제외)
            directions = []
            if not is_tsl:
                route_ports_info = route_direction_map.get(selected_route, {'first': 'PUS', 'second': 'OSA'})
                first_port = route_ports_info['first']
                second_port = route_ports_info['second']
                
                if selected_origin == '전체' and selected_destination == '전체':
                    directions = ['E', 'W']
                elif selected_origin == first_port and selected_destination == '전체':
                    directions = ['E']
                elif selected_origin == second_port and selected_destination == '전체':
                    directions = ['W']
                elif selected_origin == '전체' and selected_destination == second_port:
                    directions = ['E']
                elif selected_origin == '전체' and selected_destination == first_port:
                    directions = ['W']
                elif selected_origin == first_port and selected_destination == second_port:
                    directions = ['E']
                elif selected_origin == second_port and selected_destination == first_port:
                    directions = ['W']
                else:
                    directions = ['E', 'W']
            
            # 스케줄 조회
            all_schedules = []
            
            if is_tsl:
                # TSL: 모든 스케줄 가져오기 (port 정보 포함)
                schedule_query = f"""
                    SELECT 
                        cs.id AS schedule_id,
                        CONVERT(VARCHAR, cs.etd, 23) AS etd_date,
                        CONVERT(VARCHAR, cs.etd, 108) AS etd_time,
                        voy.route_id,
                        voy.direction,
                        ps.port_id AS departure_port_id
                    FROM coastal_schedules cs
                    LEFT JOIN proforma_schedules ps ON cs.proforma_schedule_id = ps.id
                    LEFT JOIN voyages voy ON ps.voyage_id = voy.id
                    WHERE voy.route_id = {selected_route_id}
                      AND CAST(cs.etd AS DATE) BETWEEN '{start_date}' AND '{end_date}'
                      AND cs.deleted_at IS NULL
                      AND cs.is_cruise_available = 1
                    ORDER BY cs.etd
                """
                df_schedules = pd.read_sql(schedule_query, conn_base)
            else:
                # 기타 항로: direction으로 필터링
                for direction in directions:
                    schedule_query = f"""
                        SELECT 
                            cs.id AS schedule_id,
                            CONVERT(VARCHAR, cs.etd, 23) AS etd_date,
                            CONVERT(VARCHAR, cs.etd, 108) AS etd_time,
                            voy.route_id,
                            voy.direction,
                            ps.port_id AS departure_port_id
                        FROM coastal_schedules cs
                        LEFT JOIN proforma_schedules ps ON cs.proforma_schedule_id = ps.id
                        LEFT JOIN voyages voy ON ps.voyage_id = voy.id
                        WHERE voy.route_id = {selected_route_id}
                          AND voy.direction = '{direction}'
                          AND CAST(cs.etd AS DATE) BETWEEN '{start_date}' AND '{end_date}'
                          AND cs.deleted_at IS NULL
                          AND cs.is_cruise_available = 1
                        ORDER BY cs.etd
                    """
                    df_temp = pd.read_sql(schedule_query, conn_base)
                    if not df_temp.empty:
                        all_schedules.append(df_temp)
                
                if all_schedules:
                    df_schedules = pd.concat(all_schedules, ignore_index=True)
                else:
                    df_schedules = pd.DataFrame()
            
            conn_base.close()
            
            # 날짜/시간 포맷팅 (pandas에서 처리)
            df_schedules['date'] = pd.to_datetime(df_schedules['etd_date'])
            df_schedules['date_display'] = df_schedules['date'].dt.strftime('%m-%d')
            df_schedules['weekday'] = df_schedules['date'].dt.day_name()
            weekday_ko = {
                'Monday': '월', 'Tuesday': '화', 'Wednesday': '수', 
                'Thursday': '목', 'Friday': '금', 'Saturday': '토', 'Sunday': '일'
            }
            df_schedules['weekday'] = df_schedules['weekday'].map(weekday_ko)
            # 시간 정보 추출 (HH:MM 형식)
            df_schedules['time_display'] = df_schedules['etd_time'].str[:5] if 'etd_time' in df_schedules.columns else ''
            df_schedules['date'] = df_schedules['date'].dt.date
            
            # 출발 포트 코드 추가
            if 'departure_port_id' in df_schedules.columns:
                df_schedules['departure_port'] = df_schedules['departure_port_id'].map(PORT_CODE_MAP).fillna('-')
            else:
                df_schedules['departure_port'] = '-'
            
            if df_schedules.empty:
                st.warning("해당 기간에 스케줄이 없습니다.")
            else:
                schedule_ids = ','.join(map(str, df_schedules['schedule_id'].tolist()))
                
                # route_id 목록 가져오기 (중복 제거)
                route_ids = df_schedules['route_id'].unique().tolist()
                route_ids_str = ','.join(map(str, route_ids))
                
                # TSL 필터링을 위한 arrival port 조건 준비
                # Azure SQL에서는 Cross-database 쿼리 불가 → Python에서 필터링
                tsl_arrival_filter = ""
                tsl_arrival_join = ""
                
                # 생성처 필터는 승객 분석 탭에서만 적용 (여기서는 빈 문자열)
                origin_country_filter = ""
                tsl_valid_schedule_ids = None  # TSL 도착지 필터용
                
                if is_tsl and selected_destination != '전체':
                    arrival_port_id = TSL_PORT_IDS.get(selected_destination)
                    if arrival_port_id:
                        # neohelios_base에서 schedule_id → port_id 매핑 가져오기
                        conn_base_for_tsl = pyodbc.connect(conn_string_base)
                        port_mapping_query = f"""
                            SELECT cs.id AS schedule_id, ps.port_id
                            FROM coastal_schedules cs
                            INNER JOIN proforma_schedules ps ON cs.proforma_schedule_id = ps.id
                            WHERE cs.id IN ({schedule_ids})
                        """
                        df_port_mapping = pd.read_sql(port_mapping_query, conn_base_for_tsl)
                        conn_base_for_tsl.close()
                        
                        # arrival_schedule_id가 선택한 도착 port인 티켓만 조회하기 위해
                        # 해당 port_id를 가진 schedule_id 목록 생성
                        arrival_schedule_ids = df_port_mapping[df_port_mapping['port_id'] == arrival_port_id]['schedule_id'].tolist()
                        if arrival_schedule_ids:
                            arrival_ids_str = ','.join(map(str, arrival_schedule_ids))
                            tsl_arrival_filter = f" AND t.arrival_schedule_id IN ({arrival_ids_str})"
                
                # TSL 출발지 필터 (departure_schedule_id의 port)
                tsl_departure_filter = ""
                if is_tsl and selected_origin != '전체':
                    origin_port_id = TSL_PORT_IDS.get(selected_origin)
                    if origin_port_id:
                        # departure_schedule_id의 port 확인
                        filtered_schedule_ids = df_schedules[df_schedules['departure_port_id'] == origin_port_id]['schedule_id'].tolist()
                        if filtered_schedule_ids:
                            schedule_ids = ','.join(map(str, filtered_schedule_ids))
                            # df_schedules도 필터링
                            df_schedules = df_schedules[df_schedules['departure_port_id'] == origin_port_id].copy()
                        else:
                            st.warning("선택한 출발지에 해당하는 스케줄이 없습니다.")
                            schedule_ids = ""
                
                # schedule_ids가 비어있으면 조회 중단
                if not schedule_ids:
                    st.warning("조건에 맞는 스케줄이 없습니다.")
                    st.stop()
                
                # 2. 전체 객실 수 조회 (선택한 route 기준)
                conn_cruise = pyodbc.connect(conn_string_cruise)
                total_rooms_query = f"""
                    SELECT 
                        g.code AS grade,
                        COUNT(*) AS total_rooms
                    FROM rooms r
                    JOIN grades g ON r.grade_id = g.id
                    WHERE g.route_id IN ({route_ids_str})
                      AND r.deleted_at IS NULL
                      AND g.deleted_at IS NULL
                    GROUP BY g.code
                """
                df_total_rooms = pd.read_sql(total_rooms_query, conn_cruise)
                
                # 3. 예약 현황 조회 (확정, 블록 객실/좌석 수)
                # PSMC (객실 기반): on_boarding_room_id로 객실 연결
                # PSTL/PSGR (좌석 기반): grade_price_detail_by_age_group_id로 등급 연결
                # 확정: is_temporary=0, 블록: is_temporary=1
                # REFUND 상태는 취소 티켓이므로 제외!
                
                is_seat_based = selected_vessel in seat_based_vessels
                
                if is_seat_based:
                    # PSTL/PSGR: 좌석 기반 - grade_price_detail_by_age_group_id 경로로 등급 조회
                    # 블록 티켓은 on_boarding_room_id가 NULL이므로 이 경로 사용
                    booking_query = f"""
                        SELECT 
                            t.departure_schedule_id AS schedule_id,
                            g.code AS grade,
                            COUNT(CASE WHEN t.is_temporary = 0 AND t.status NOT LIKE 'REFUND%' THEN 1 END) AS confirmed_rooms,
                            COUNT(CASE WHEN t.is_temporary = 1 AND t.status NOT LIKE 'REFUND%' THEN 1 END) AS blocked_rooms
                        FROM tickets t
                        LEFT JOIN grade_price_detail_by_age_groups gpdag ON t.grade_price_detail_by_age_group_id = gpdag.id
                        LEFT JOIN grade_price_details gpd ON gpdag.grade_price_detail_id = gpd.id
                        LEFT JOIN grade_prices gp ON gpd.grade_price_id = gp.id
                        LEFT JOIN grades g ON gp.grade_id = g.id
                        WHERE t.departure_schedule_id IN ({schedule_ids})
                          AND t.deleted_at IS NULL
                          AND t.status NOT LIKE 'REFUND%'
                          AND g.code IS NOT NULL
                          {tsl_arrival_filter}
                          {origin_country_filter}
                        GROUP BY t.departure_schedule_id, g.code
                    """
                else:
                    # PSMC: 객실 기반 - on_boarding_room_id로 객실 연결
                    booking_query = f"""
                        WITH room_status AS (
                            SELECT 
                                t.departure_schedule_id,
                                t.on_boarding_room_id,
                                g.code AS grade,
                                MAX(CASE 
                                    WHEN t.is_temporary = 0 
                                         AND t.status NOT LIKE 'REFUND%'
                                    THEN 1 
                                    ELSE 0 
                                END) AS has_confirmed,
                                MAX(CASE 
                                    WHEN t.is_temporary = 1 
                                         AND t.status NOT LIKE 'REFUND%'
                                    THEN 1 
                                    ELSE 0 
                                END) AS has_blocked
                            FROM tickets t
                            INNER JOIN rooms r ON t.on_boarding_room_id = r.id
                            INNER JOIN grades g ON r.grade_id = g.id
                            {tsl_arrival_join}
                            WHERE t.departure_schedule_id IN ({schedule_ids})
                              AND t.deleted_at IS NULL
                              AND r.deleted_at IS NULL
                              AND g.deleted_at IS NULL
                              AND t.on_boarding_room_id IS NOT NULL
                              AND t.status NOT LIKE 'REFUND%'
                              {tsl_arrival_filter}
                              {origin_country_filter}
                            GROUP BY t.departure_schedule_id, t.on_boarding_room_id, g.code
                        )
                        SELECT 
                            departure_schedule_id AS schedule_id,
                            grade,
                            COUNT(CASE WHEN has_confirmed = 1 THEN 1 END) AS confirmed_rooms,
                            COUNT(CASE WHEN has_confirmed = 0 AND has_blocked = 1 THEN 1 END) AS blocked_rooms
                        FROM room_status
                        WHERE grade IS NOT NULL
                        GROUP BY departure_schedule_id, grade
                    """
                df_bookings = pd.read_sql(booking_query, conn_cruise)
                
                # 3-1. 승객 수 조회 (티켓 수 기반)
                # PSMC: 확정=티켓수, 블록=MIN(블록티켓수, 정원)
                # PSTL/PSGR: 1좌석=1승객이므로 좌석 수와 동일
                
                if is_seat_based:
                    # PSTL/PSGR: 좌석 기반 - 좌석 수 = 승객 수
                    passenger_query = f"""
                        SELECT 
                            t.departure_schedule_id AS schedule_id,
                            g.code AS grade,
                            COUNT(CASE WHEN t.is_temporary = 0 AND t.status NOT LIKE 'REFUND%' THEN 1 END) AS confirmed_passengers,
                            COUNT(CASE WHEN t.is_temporary = 1 AND t.status NOT LIKE 'REFUND%' THEN 1 END) AS blocked_passengers
                        FROM tickets t
                        LEFT JOIN grade_price_detail_by_age_groups gpdag ON t.grade_price_detail_by_age_group_id = gpdag.id
                        LEFT JOIN grade_price_details gpd ON gpdag.grade_price_detail_id = gpd.id
                        LEFT JOIN grade_prices gp ON gpd.grade_price_id = gp.id
                        LEFT JOIN grades g ON gp.grade_id = g.id
                        WHERE t.departure_schedule_id IN ({schedule_ids})
                          AND t.deleted_at IS NULL
                          AND t.status NOT LIKE 'REFUND%'
                          AND g.code IS NOT NULL
                          {tsl_arrival_filter}
                          {origin_country_filter}
                        GROUP BY t.departure_schedule_id, g.code
                    """
                else:
                    # PSMC: 객실 기반 - 정원 제한 적용
                    # 확정: grade_price_detail_by_age_group_id 경로로 등급 조회 (on_boarding_room_id 조건 없음)
                    # 블록: rooms 기반으로 정원 제한 적용
                    passenger_query = f"""
                        WITH confirmed_count AS (
                            SELECT 
                                t.departure_schedule_id,
                                g.code AS grade,
                                COUNT(*) AS confirmed_passengers
                            FROM tickets t
                            LEFT JOIN grade_price_detail_by_age_groups gpdag ON t.grade_price_detail_by_age_group_id = gpdag.id
                            LEFT JOIN grade_price_details gpd ON gpdag.grade_price_detail_id = gpd.id
                            LEFT JOIN grade_prices gp ON gpd.grade_price_id = gp.id
                            LEFT JOIN grades g ON gp.grade_id = g.id
                            {tsl_arrival_join}
                            WHERE t.departure_schedule_id IN ({schedule_ids})
                              AND t.deleted_at IS NULL
                              AND t.is_temporary = 0
                              AND t.status NOT LIKE 'REFUND%'
                              AND g.code IS NOT NULL
                              {tsl_arrival_filter}
                              {origin_country_filter}
                            GROUP BY t.departure_schedule_id, g.code
                        ),
                        room_blocked AS (
                            SELECT 
                                t.departure_schedule_id,
                                t.on_boarding_room_id,
                                g.code AS grade,
                                COUNT(*) AS blocked_tickets,
                                CASE 
                                    WHEN g.code IN ('OR', 'BS', 'PR') THEN 2
                                    WHEN g.code = 'RS' THEN 3
                                    WHEN g.code IN ('IC', 'OC', 'DA') THEN 4
                                    WHEN g.code = 'GR' THEN 8
                                    ELSE 2
                                END AS capacity
                            FROM tickets t
                            INNER JOIN rooms r ON t.on_boarding_room_id = r.id
                            INNER JOIN grades g ON r.grade_id = g.id
                            {tsl_arrival_join}
                            WHERE t.departure_schedule_id IN ({schedule_ids})
                              AND t.deleted_at IS NULL
                              AND r.deleted_at IS NULL
                              AND g.deleted_at IS NULL
                              AND t.on_boarding_room_id IS NOT NULL
                              AND t.is_temporary = 1
                              AND t.status NOT LIKE 'REFUND%'
                              {tsl_arrival_filter}
                              {origin_country_filter}
                            GROUP BY t.departure_schedule_id, t.on_boarding_room_id, g.code
                        ),
                        blocked_count AS (
                            SELECT 
                                departure_schedule_id,
                                grade,
                                SUM(CASE WHEN blocked_tickets <= capacity THEN blocked_tickets ELSE capacity END) AS blocked_passengers
                            FROM room_blocked
                            GROUP BY departure_schedule_id, grade
                        )
                        SELECT 
                            COALESCE(c.departure_schedule_id, b.departure_schedule_id) AS schedule_id,
                            COALESCE(c.grade, b.grade) AS grade,
                            COALESCE(c.confirmed_passengers, 0) AS confirmed_passengers,
                            COALESCE(b.blocked_passengers, 0) AS blocked_passengers
                        FROM confirmed_count c
                        FULL OUTER JOIN blocked_count b 
                            ON c.departure_schedule_id = b.departure_schedule_id 
                            AND c.grade = b.grade
                    """
                df_passengers = pd.read_sql(passenger_query, conn_cruise)
                
                # 3-2. 객실/좌석별 상세 정보 조회 (모달용)
                if is_seat_based:
                    # PSTL/PSGR: 좌석 기반 - 티켓별로 조회
                    # 확정 좌석은 room_number, 블록 좌석은 티켓ID 일부 사용
                    room_details_query = f"""
                        SELECT 
                            t.departure_schedule_id AS schedule_id,
                            g.code AS grade,
                            COALESCE(r.room_number, CAST(t.id AS VARCHAR(20))) AS room_no,
                            CASE 
                                WHEN t.is_temporary = 0 THEN 'confirmed'
                                WHEN t.is_temporary = 1 THEN 'blocked'
                            END AS status
                        FROM tickets t
                        LEFT JOIN rooms r ON t.on_boarding_room_id = r.id
                        LEFT JOIN grade_price_detail_by_age_groups gpdag ON t.grade_price_detail_by_age_group_id = gpdag.id
                        LEFT JOIN grade_price_details gpd ON gpdag.grade_price_detail_id = gpd.id
                        LEFT JOIN grade_prices gp ON gpd.grade_price_id = gp.id
                        LEFT JOIN grades g ON gp.grade_id = g.id
                        WHERE t.departure_schedule_id IN ({schedule_ids})
                          AND t.deleted_at IS NULL
                          AND t.status NOT LIKE 'REFUND%'
                          AND g.code IS NOT NULL
                          {tsl_arrival_filter}
                          {origin_country_filter}
                        ORDER BY schedule_id, grade, room_no
                    """
                else:
                    # PSMC: 객실 기반
                    room_details_query = f"""
                        WITH room_status AS (
                            SELECT 
                                t.departure_schedule_id,
                                t.on_boarding_room_id,
                                r.room_number,
                                g.code AS grade,
                                MAX(CASE 
                                    WHEN t.is_temporary = 0 
                                         AND t.status NOT LIKE 'REFUND%'
                                    THEN 1 
                                    ELSE 0 
                                END) AS has_confirmed,
                                MAX(CASE 
                                    WHEN t.is_temporary = 1 
                                         AND t.status NOT LIKE 'REFUND%'
                                    THEN 1 
                                    ELSE 0 
                                END) AS has_blocked
                            FROM tickets t
                            INNER JOIN rooms r ON t.on_boarding_room_id = r.id
                            INNER JOIN grades g ON r.grade_id = g.id
                            {tsl_arrival_join}
                            WHERE t.departure_schedule_id IN ({schedule_ids})
                              AND t.deleted_at IS NULL
                              AND r.deleted_at IS NULL
                              AND g.deleted_at IS NULL
                              AND t.on_boarding_room_id IS NOT NULL
                              AND t.status NOT LIKE 'REFUND%'
                              {tsl_arrival_filter}
                              {origin_country_filter}
                            GROUP BY t.departure_schedule_id, t.on_boarding_room_id, r.room_number, g.code
                        )
                        SELECT 
                            departure_schedule_id AS schedule_id,
                            grade,
                            room_number AS room_no,
                            CASE 
                                WHEN has_confirmed = 1 THEN 'confirmed'
                                WHEN has_blocked = 1 THEN 'blocked'
                            END AS status
                        FROM room_status
                        WHERE grade IS NOT NULL
                        ORDER BY schedule_id, grade, room_number
                    """
                df_room_details = pd.read_sql(room_details_query, conn_cruise)
                
                # 공실 목록도 추가 (전체 객실에서 예약된 객실 제외)
                # PSTL/PSGR은 좌석이 수백 개라 공실 목록 표시 안함
                if is_seat_based:
                    # 빈 DataFrame 생성
                    df_vacant_rooms = pd.DataFrame(columns=['schedule_id', 'grade', 'room_no', 'status'])
                else:
                    # PSMC: 객실 기반 - 공실 목록 조회
                    schedule_values = ','.join([f"({sid})" for sid in df_schedules['schedule_id'].tolist()])
                    
                    vacant_rooms_query = f"""
                        SELECT 
                            cs.schedule_id,
                            g.code AS grade,
                            r.room_number AS room_no
                        FROM rooms r
                        INNER JOIN grades g ON r.grade_id = g.id
                        CROSS JOIN (
                            SELECT * FROM (VALUES {schedule_values}) AS t(schedule_id)
                        ) cs
                        WHERE g.route_id IN ({route_ids_str})
                          AND r.deleted_at IS NULL
                          AND g.deleted_at IS NULL
                          AND NOT EXISTS (
                              SELECT 1 FROM tickets t
                              WHERE t.on_boarding_room_id = r.id
                                AND t.departure_schedule_id = cs.schedule_id
                                AND t.deleted_at IS NULL
                                AND t.status NOT LIKE 'REFUND%'
                          )
                        ORDER BY schedule_id, grade, room_number
                    """
                    df_vacant_rooms = pd.read_sql(vacant_rooms_query, conn_cruise)
                    df_vacant_rooms['status'] = 'vacant'
                
                # 4. 승객 분석 데이터 조회 (확정 승객만)
                # 성별, 국적, 연령대 분석용
                # birth_day는 datetimeoffset 타입이라 CONVERT로 date로 변환
                # arrival_schedule_id 추가 (생성처별 분석용)
                passenger_analysis_query = f"""
                    SELECT 
                        t.departure_schedule_id AS schedule_id,
                        t.arrival_schedule_id,
                        p.sex,
                        p.nationality,
                        CONVERT(date, p.birth_day) AS birth_day,
                        ISNULL(t.is_issued, 0) AS is_issued,
                        t.ticket_number
                    FROM tickets t
                    INNER JOIN reservation_passengers rp ON t.reservation_passenger_id = rp.id
                    INNER JOIN passengers p ON rp.passenger_id = p.id
                    WHERE t.departure_schedule_id IN ({schedule_ids})
                      AND t.is_temporary = 0
                      AND t.deleted_at IS NULL
                      AND t.status NOT LIKE 'REFUND%'
                      AND rp.deleted_at IS NULL
                      AND p.deleted_at IS NULL
                      {tsl_arrival_filter}
                """
                df_passenger_analysis = pd.read_sql(passenger_analysis_query, conn_cruise)
                
                conn_cruise.close()
                
                # 4. 데이터 병합 및 공실 계산
                # 출발/도착 포트 계산
                route_ports_info = route_direction_map.get(selected_route, {'first': '-', 'second': '-'})
                first_port = route_ports_info.get('first', '-')
                second_port = route_ports_info.get('second', '-')
                
                # 모든 스케줄 x 모든 등급 조합 생성
                all_combinations = []
                for _, schedule in df_schedules.iterrows():
                    direction = schedule.get('direction', '')
                    dep_port = schedule.get('departure_port', '-')
                    # E방향: 첫번째→두번째, W방향: 두번째→첫번째
                    arr_port = second_port if direction == 'E' else (first_port if direction == 'W' else '-')
                    
                    for _, grade_info in df_total_rooms.iterrows():
                        all_combinations.append({
                            'schedule_id': schedule['schedule_id'],
                            'date': schedule['date'],
                            'date_display': schedule['date_display'],
                            'weekday': schedule['weekday'],
                            'time_display': schedule.get('time_display', ''),
                            'direction': direction,
                            'departure_port': dep_port,
                            'arrival_port': arr_port,
                            'grade': grade_info['grade'],
                            'total_rooms': grade_info['total_rooms']
                        })
                
                df_all = pd.DataFrame(all_combinations)
                
                # 예약 현황 병합
                df_result = df_all.merge(df_bookings, on=['schedule_id', 'grade'], how='left')
                df_result['confirmed_rooms'] = df_result['confirmed_rooms'].fillna(0).astype(int)
                df_result['blocked_rooms'] = df_result['blocked_rooms'].fillna(0).astype(int)
                df_result['total_rooms'] = df_result['total_rooms'].astype(int)
                
                # 공실 계산
                df_result['vacant_rooms'] = df_result['total_rooms'] - df_result['confirmed_rooms'] - df_result['blocked_rooms']
                df_result['vacant_rooms'] = df_result['vacant_rooms'].clip(lower=0).astype(int)
                
                # 5. 스케줄별 총계 계산 (하루에 여러 편 운항 고려)
                df_totals = df_result.groupby(['schedule_id', 'date', 'date_display', 'weekday', 'time_display', 'direction', 'departure_port', 'arrival_port']).agg({
                    'confirmed_rooms': 'sum',
                    'blocked_rooms': 'sum',
                    'vacant_rooms': 'sum'
                }).reset_index()
                df_totals['grade'] = '총계'
                
                # 6. 총계와 등급별 데이터 합치기
                df_with_totals = pd.concat([df_totals, df_result], ignore_index=True)
                
                # 7. 날짜+시간 표시 형식 (하루에 여러 편이면 시간 표시)
                # 같은 날짜에 여러 스케줄이 있는지 확인
                schedules_per_date = df_schedules.groupby('date').size()
                has_multiple_schedules = (schedules_per_date > 1).any()
                
                if has_multiple_schedules:
                    # 시간 표시 포함
                    df_with_totals['날짜'] = df_with_totals['date_display'] + ' ' + df_with_totals['time_display'] + ' (' + df_with_totals['weekday'] + ')'
                else:
                    # 기존 방식 (날짜만)
                    df_with_totals['날짜'] = df_with_totals['date_display'] + ' (' + df_with_totals['weekday'] + ')'
                
                # 8. 등급 순서 정의 (선박/항로별로 다름)
                # PSMC (route 1-4): OR, PR, RS, BS, OC, IC, DA
                # PSTL (route 5): PRM, ECM
                # PSGR (route 6-11): FC, BUS, STA
                if selected_vessel == 'PSMC':
                    grade_order = ['총계', 'OR', 'PR', 'RS', 'BS', 'OC', 'IC', 'DA']
                elif selected_vessel == 'PSTL':
                    grade_order = ['총계', 'PRM', 'ECM']
                else:  # PSGR
                    grade_order = ['총계', 'FC', 'BUS', 'STA']
                
                # DB에 있는 등급만 필터링
                existing_grades = [g for g in grade_order if g in df_with_totals['grade'].unique()]
                # grade_order에 없는 등급도 추가 (혹시 새로운 등급이 있을 경우)
                for g in df_with_totals['grade'].unique():
                    if g not in existing_grades and g != '총계':
                        existing_grades.append(g)
                
                # 9. 스케줄별로 한 행씩 구성 (하루에 여러 편 운항 고려)
                result_rows = []
                # schedule_id 순으로 정렬 (날짜+시간 순)
                for schedule_id in df_schedules.sort_values(['date', 'etd_time'])['schedule_id'].unique():
                    schedule_data = df_with_totals[df_with_totals['schedule_id'] == schedule_id]
                    if schedule_data.empty:
                        continue
                    
                    # 총계의 확정+블록이 0이면 스킵 (예약이 하나도 없는 스케줄)
                    total_data = schedule_data[schedule_data['grade'] == '총계']
                    if not total_data.empty:
                        total_confirmed = int(total_data['confirmed_rooms'].iloc[0])
                        total_blocked = int(total_data['blocked_rooms'].iloc[0])
                        if total_confirmed == 0 and total_blocked == 0:
                            continue  # 예약 없는 스케줄 숨기기
                    
                    row = {
                        '날짜': schedule_data['날짜'].iloc[0],
                        'schedule_id': schedule_id,
                        'date_raw': str(schedule_data['date'].iloc[0]),
                        'departure_port': schedule_data['departure_port'].iloc[0] if 'departure_port' in schedule_data.columns else '-',
                        'arrival_port': schedule_data['arrival_port'].iloc[0] if 'arrival_port' in schedule_data.columns else '-'
                    }
                    
                    for grade in existing_grades:
                        grade_data = schedule_data[schedule_data['grade'] == grade]
                        if not grade_data.empty:
                            row[f'{grade}_확정'] = int(grade_data['confirmed_rooms'].iloc[0])
                            row[f'{grade}_블록'] = int(grade_data['blocked_rooms'].iloc[0])
                            row[f'{grade}_공실'] = int(grade_data['vacant_rooms'].iloc[0])
                        else:
                            row[f'{grade}_확정'] = 0
                            row[f'{grade}_블록'] = 0
                            row[f'{grade}_공실'] = 0
                    
                    result_rows.append(row)
                
                # 10. DataFrame 생성
                final_df = pd.DataFrame(result_rows)
                
                # 11. 컬럼 순서 정리 (schedule_id, date_raw, departure_port, arrival_port 유지)
                ordered_cols = ['날짜', 'schedule_id', 'date_raw', 'departure_port', 'arrival_port']
                for grade in existing_grades:
                    ordered_cols.extend([f'{grade}_확정', f'{grade}_블록', f'{grade}_공실'])
                
                final_df = final_df[ordered_cols]
                
                # 12. 깔끔한 테이블 생성
                st.markdown(f'<div style="background: #F3F7F9; padding: 12px 20px; border-radius: 5px; color: #232A5E; font-weight: 500; font-size: 14px; margin: 16px 0; border-left: 4px solid #232A5E; font-family: Noto Sans KR, sans-serif;">✓ {len(df_schedules)}개 스케줄 조회 완료</div>', unsafe_allow_html=True)
                
                # HTML 테이블 생성 (피그마 디자인 시스템)
                html_table = '<div class="responsive-table-container"><table style="width:100%; border-collapse: collapse; background: #FFFFFF; font-family: Noto Sans KR, sans-serif;">'
                
                # 헤더 1행: 등급명 - NEOHELIOS 디자인 시스템
                html_table += '<thead><tr><th rowspan="2" class="sticky-date-header" style="background: #232A5E; color: #FAFCFE; padding: 12px 10px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; letter-spacing: -0.5px; text-align: center;">날짜</th>'
                html_table += '<th rowspan="2" style="background: #232A5E; color: #FAFCFE; padding: 12px 8px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; letter-spacing: -0.5px; text-align: center;">출발</th>'
                html_table += '<th rowspan="2" style="background: #232A5E; color: #FAFCFE; padding: 12px 8px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; letter-spacing: -0.5px; text-align: center;">도착</th>'
                for idx, grade in enumerate(existing_grades):
                    if grade == '총계':
                        bg_color = '#1a2148'
                    else:
                        bg_color = '#232A5E'
                    
                    is_last_grade = (idx == len(existing_grades) - 1)
                    border_right = '1px solid #3a4a7e' if is_last_grade else '1px solid #3a4a7e'
                    
                    html_table += f'<th colspan="3" style="background: {bg_color}; color: #FAFCFE; padding: 12px 10px; border: none; border-right: {border_right}; font-weight: 700; font-size: 12px; letter-spacing: -0.5px; text-align: center;">{grade}</th>'
                html_table += '</tr>'
                
                # 헤더 2행: 확정/블록/공실
                html_table += '<tr>'
                for idx, grade in enumerate(existing_grades):
                    is_last_grade = (idx == len(existing_grades) - 1)
                    grade_separator = '1px solid #DAE0E3' if is_last_grade else '1px solid #c8d0d4'
                    
                    html_table += '<th style="background: #F3F7F9; color: #232A5E; text-align: center; padding: 10px 8px; font-weight: 500; border: none; border-right: 1px solid #DAE0E3; border-bottom: 1px solid #DAE0E3; font-size: 12px; letter-spacing: -0.5px;">확정</th>'
                    html_table += '<th style="background: #F3F7F9; color: #232A5E; text-align: center; padding: 10px 8px; font-weight: 500; border: none; border-right: 1px solid #DAE0E3; border-bottom: 1px solid #DAE0E3; font-size: 12px; letter-spacing: -0.5px;">블록</th>'
                    html_table += f'<th style="background: #FFFBEB; color: #232A5E; text-align: center; padding: 10px 8px; font-weight: 500; border: none; border-right: {grade_separator}; border-bottom: 1px solid #DAE0E3; font-size: 12px; letter-spacing: -0.5px;">공실</th>'
                html_table += '</tr></thead>'
                
                # 선박 정보 (항로에 따라 고정)
                vessel_name = selected_vessel
                
                # 바디 - NEOHELIOS 디자인 시스템
                html_table += '<tbody>'
                for idx, row in final_df.iterrows():
                    # 교차 행 배경
                    row_bg = '#FFFFFF' if idx % 2 == 0 else '#F9FAFB'
                    # schedule_id NaN 처리
                    schedule_id_raw = row.get('schedule_id', 0)
                    schedule_id = int(schedule_id_raw) if pd.notna(schedule_id_raw) else 0
                    date_raw = row.get('date_raw', '')
                    
                    dep_port = row.get('departure_port', '-')
                    arr_port = row.get('arrival_port', '-')
                    
                    html_table += '<tr style="border-bottom: 1px solid #DAE0E3; transition: background 0.15s ease;">'
                    html_table += f'<td class="sticky-date-cell" style="background: {row_bg}; color: #0E0E2C; font-weight: 500; padding: 10px; border: none; border-right: 1px solid #DAE0E3; font-size: 14px; letter-spacing: -0.5px; text-align: center;">{row["날짜"]}</td>'
                    html_table += f'<td style="background: {row_bg}; color: #0E0E2C; font-weight: 500; padding: 10px 8px; border: none; border-right: 1px solid #DAE0E3; font-size: 14px; letter-spacing: -0.5px; text-align: center;">{dep_port}</td>'
                    html_table += f'<td style="background: {row_bg}; color: #0E0E2C; font-weight: 500; padding: 10px 8px; border: none; border-right: 1px solid #DAE0E3; font-size: 14px; letter-spacing: -0.5px; text-align: center;">{arr_port}</td>'
                    
                    for grade_idx, grade in enumerate(existing_grades):
                        confirmed = int(row.get(f'{grade}_확정', 0))
                        blocked = int(row.get(f'{grade}_블록', 0))
                        vacant = int(row.get(f'{grade}_공실', 0))
                        date_display = row['날짜']
                        
                        # 등급 간 구분선
                        is_last_grade = (grade_idx == len(existing_grades) - 1)
                        grade_separator = '1px solid #DAE0E3' if is_last_grade else '1px solid #c8d0d4'
                        
                        # 클릭 가능 여부 (총계는 클릭 불가, schedule_id가 없으면 불가)
                        is_clickable = grade != '총계' and schedule_id > 0
                        
                        if is_clickable:
                            # JavaScript onclick으로 모달 표시
                            confirmed_link = f'<span onclick="openRoomModal({schedule_id}, \'{date_display}\', \'{grade}\', \'confirmed\')" style="cursor: pointer; display: block;" title="클릭하여 상세보기">{confirmed}</span>'
                            
                            # PSTL/PSGR 좌석 기반: 블록과 공실은 클릭 불가
                            if is_seat_based:
                                blocked_link = str(blocked)
                                vacant_link = str(vacant)
                            else:
                                blocked_link = f'<span onclick="openRoomModal({schedule_id}, \'{date_display}\', \'{grade}\', \'blocked\')" style="cursor: pointer; display: block;" title="클릭하여 상세보기">{blocked}</span>'
                                vacant_link = f'<span onclick="openRoomModal({schedule_id}, \'{date_display}\', \'{grade}\', \'vacant\')" style="cursor: pointer; display: block;" title="클릭하여 상세보기">{vacant}</span>'
                            cell_class = 'class="clickable-cell"'
                        else:
                            confirmed_link = str(confirmed)
                            blocked_link = str(blocked)
                            vacant_link = str(vacant)
                            cell_class = ''
                        
                        # 확정: Primary 색상
                        html_table += f'<td {cell_class} style="background: {row_bg}; color: #0E0E2C; text-align: center; padding: 10px; font-weight: 500; border: none; border-right: 1px solid #DAE0E3; font-size: 14px;">{confirmed_link}</td>'
                        
                        # 블록: 그레이 톤
                        html_table += f'<td {cell_class} style="background: {row_bg}; color: #88949C; text-align: center; padding: 10px; font-weight: 400; border: none; border-right: 1px solid #DAE0E3; font-size: 14px;">{blocked_link}</td>'
                        
                        # 공실: 옅은 노란색 배경, 0이면 예약불가 강조
                        if vacant == 0:
                            # 예약 불가 - Alert Red 강조
                            vacant_style = f'background: #FEF2F2; color: #EA3336; text-align: center; padding: 10px; font-weight: 700; border: none; border-right: {grade_separator}; border-left: 3px solid #EA3336; font-size: 14px;'
                        else:
                            # 일반 공실 - 옅은 노란색 배경
                            yellow_bg = '#FFFBEB' if row_bg == '#FFFFFF' else '#FEF9E7'
                            vacant_style = f'background: {yellow_bg}; color: #436CFC; text-align: center; padding: 10px; font-weight: 500; border: none; border-right: {grade_separator}; font-size: 14px;'
                        
                        html_table += f'<td {cell_class} style="{vacant_style}">{vacant_link}</td>'
                    
                    html_table += '</tr>'
                html_table += '</tbody></table></div>'
                
                # ========== 승객 수 기반 테이블 생성 ==========
                # 등급별 정원 정의 (OR,BS,PR=2명, RS=3명, IC,OC,DA=4명, GR=8명)
                grade_capacity = {
                    'OR': 2, 'BS': 2, 'PR': 2, 'RS': 3,
                    'IC': 4, 'OC': 4, 'DA': 4,
                    'GR': 8,
                    'PRM': 1, 'ECM': 1,  # PSTL 좌석
                    'FC': 1, 'BUS': 1, 'STA': 1  # PSGR 좌석
                }
                
                # 승객 데이터 병합
                df_pass_result = df_all[['schedule_id', 'date', 'date_display', 'weekday', 'grade', 'total_rooms']].merge(
                    df_passengers, on=['schedule_id', 'grade'], how='left'
                )
                df_pass_result['confirmed_passengers'] = df_pass_result['confirmed_passengers'].fillna(0).astype(int)
                df_pass_result['blocked_passengers'] = df_pass_result['blocked_passengers'].fillna(0).astype(int)
                
                # 등급별 총 정원 계산 (정원 × 객실수)
                df_pass_result['capacity'] = df_pass_result['grade'].map(grade_capacity).fillna(2).astype(int)
                df_pass_result['total_capacity'] = df_pass_result['total_rooms'] * df_pass_result['capacity']
                
                # 잔여 계산 (총 정원 - 확정 - 블록)
                df_pass_result['remaining_passengers'] = (
                    df_pass_result['total_capacity'] - 
                    df_pass_result['confirmed_passengers'] - 
                    df_pass_result['blocked_passengers']
                ).clip(lower=0).astype(int)
                
                # 승객 총계 계산 (스케줄별)
                df_pass_totals = df_pass_result.groupby(['schedule_id', 'date', 'date_display', 'weekday']).agg({
                    'confirmed_passengers': 'sum',
                    'blocked_passengers': 'sum',
                    'remaining_passengers': 'sum'
                }).reset_index()
                df_pass_totals['grade'] = '총계'
                
                # 승객 총계와 등급별 데이터 합치기 (remaining_passengers 컬럼 포함)
                df_pass_with_totals = pd.concat([df_pass_totals, df_pass_result[['schedule_id', 'date', 'date_display', 'weekday', 'grade', 'confirmed_passengers', 'blocked_passengers', 'remaining_passengers']]], ignore_index=True)
                
                # 날짜+시간 표시 (객실 탭과 동일)
                if has_multiple_schedules:
                    # 스케줄별 시간 정보 가져오기
                    schedule_time_map = df_schedules.set_index('schedule_id')['time_display'].to_dict()
                    df_pass_with_totals['time_display'] = df_pass_with_totals['schedule_id'].map(schedule_time_map).fillna('')
                    df_pass_with_totals['날짜'] = df_pass_with_totals['date_display'] + ' ' + df_pass_with_totals['time_display'] + ' (' + df_pass_with_totals['weekday'] + ')'
                else:
                    df_pass_with_totals['날짜'] = df_pass_with_totals['date_display'] + ' (' + df_pass_with_totals['weekday'] + ')'
                
                # 승객 스케줄별로 한 행씩 구성 (공실 포함)
                # 스케줄별 출발/도착 포트 매핑
                schedule_dep_port_map = df_schedules.set_index('schedule_id')['departure_port'].to_dict() if 'departure_port' in df_schedules.columns else {}
                schedule_direction_map = df_schedules.set_index('schedule_id')['direction'].to_dict() if 'direction' in df_schedules.columns else {}
                
                pass_result_rows = []
                for schedule_id in df_schedules.sort_values(['date', 'etd_time'])['schedule_id'].unique():
                    schedule_data = df_pass_with_totals[df_pass_with_totals['schedule_id'] == schedule_id]
                    # 같은 스케줄의 객실 공실 정보 가져오기
                    room_schedule_data = df_with_totals[df_with_totals['schedule_id'] == schedule_id]
                    
                    if schedule_data.empty:
                        continue
                    
                    # 총계의 확정+블록이 0이면 스킵 (예약이 하나도 없는 스케줄)
                    total_data = schedule_data[schedule_data['grade'] == '총계']
                    if not total_data.empty:
                        total_confirmed = int(total_data['confirmed_passengers'].iloc[0])
                        total_blocked = int(total_data['blocked_passengers'].iloc[0])
                        if total_confirmed == 0 and total_blocked == 0:
                            continue  # 예약 없는 스케줄 숨기기
                    
                    # 출발/도착 포트 계산
                    dep_port = schedule_dep_port_map.get(schedule_id, '-')
                    direction = schedule_direction_map.get(schedule_id, '')
                    arr_port = second_port if direction == 'E' else (first_port if direction == 'W' else '-')
                    
                    row = {
                        '날짜': schedule_data['날짜'].iloc[0],
                        'schedule_id': schedule_id,
                        'date_raw': str(schedule_data['date'].iloc[0]),
                        'departure_port': dep_port,
                        'arrival_port': arr_port
                    }
                    
                    for grade in existing_grades:
                        grade_data = schedule_data[schedule_data['grade'] == grade]
                        
                        if not grade_data.empty:
                            row[f'{grade}_확정'] = int(grade_data['confirmed_passengers'].iloc[0])
                            row[f'{grade}_블록'] = int(grade_data['blocked_passengers'].iloc[0])
                            row[f'{grade}_잔여'] = int(grade_data['remaining_passengers'].iloc[0])
                        else:
                            row[f'{grade}_확정'] = 0
                            row[f'{grade}_블록'] = 0
                            row[f'{grade}_잔여'] = 0
                    
                    pass_result_rows.append(row)
                
                final_df_passengers = pd.DataFrame(pass_result_rows)
                
                # 객실 상세 데이터 병합 (확정/블록 + 공실)
                # schedule_id 타입 통일 (정수형)
                df_room_details['schedule_id'] = df_room_details['schedule_id'].astype(int)
                df_vacant_rooms['schedule_id'] = df_vacant_rooms['schedule_id'].astype(int)
                df_all_room_details = pd.concat([df_room_details, df_vacant_rooms], ignore_index=True)
                
                # 조회 결과를 session_state에 저장
                st.session_state.query_result = {
                    'html_table': html_table,
                    'final_df': final_df,
                    'final_df_passengers': final_df_passengers,
                    'existing_grades': existing_grades,
                    'start_date': str(start_date),
                    'end_date': str(end_date),
                    'vessel_name': vessel_name,
                    'room_details': df_all_room_details.to_dict('records'),  # 모달용 데이터
                    'is_seat_based': is_seat_based,  # PSTL/PSGR 좌석 기반 여부
                    'passenger_analysis': df_passenger_analysis,  # 승객 분석 데이터
                    'schedules': df_schedules  # 스케줄 데이터 (생성처별 분석용)
                }
                
                st.success("조회 완료")
                
        except Exception as e:
            st.markdown(f'<div style="background: #ffebee; border-left: 3px solid #d32f2f; padding: 15px; border-radius: 4px; color: #d32f2f; font-weight: 500; margin: 20px 0;">오류 발생: {str(e)}</div>', unsafe_allow_html=True)
            st.code(str(e))

# 조회 결과 표시 (조회 버튼과 독립적으로)
if 'query_result' in st.session_state:
    result = st.session_state.query_result
    
    # 선박에 따라 탭 이름 결정 (좌석 기반 vs 객실 기반)
    vessel_name = result.get('vessel_name', 'PSMC')
    is_seat_based = vessel_name in ['PSTL', 'PSGR']
    tab1_name = "좌석" if is_seat_based else "객실"
    
    # 엑셀 다운로드 버튼용 데이터 준비 (탭 위에 배치하기 위해 미리 생성)
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    final_df = result['final_df']
    passenger_final_df = result['final_df_passengers']
    existing_grades = result['existing_grades']
    start_date = result['start_date']
    end_date = result['end_date']
    
    # 엑셀 워크북 생성
    wb = Workbook()
    
    # 시트 1: 객실
    ws = wb.active
    ws.title = '객실'
    current_col = 1
    ws.cell(1, current_col, '날짜')
    ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
    current_col += 1
    ws.cell(1, current_col, '출발')
    ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
    current_col += 1
    ws.cell(1, current_col, '도착')
    ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
    current_col += 1
    for grade in existing_grades:
        ws.cell(1, current_col, grade)
        ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 2)
        current_col += 3
    current_col = 4
    for grade in existing_grades:
        ws.cell(2, current_col, '확정')
        ws.cell(2, current_col + 1, '블록')
        ws.cell(2, current_col + 2, '공실')
        current_col += 3
    for row_idx, row in final_df.iterrows():
        excel_row = row_idx + 3
        current_col = 1
        ws.cell(excel_row, current_col, row['날짜'])
        current_col += 1
        ws.cell(excel_row, current_col, row.get('departure_port', '-'))
        current_col += 1
        ws.cell(excel_row, current_col, row.get('arrival_port', '-'))
        current_col += 1
        for grade in existing_grades:
            ws.cell(excel_row, current_col, int(row.get(f'{grade}_확정', 0)))
            ws.cell(excel_row, current_col + 1, int(row.get(f'{grade}_블록', 0)))
            ws.cell(excel_row, current_col + 2, int(row.get(f'{grade}_공실', 0)))
            current_col += 3
    
    # 시트 1 스타일링
    header_fill = PatternFill(start_color='0a0a0a', end_color='0a0a0a', fill_type='solid')
    header_font = Font(color='FFFFFF', size=12, bold=True)
    subheader_fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
    subheader_font = Font(color='6b6b6b', size=11, bold=True)
    yellow_fill = PatternFill(start_color='fffef5', end_color='fffef5', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='e0e0e0'),
        right=Side(style='thin', color='e0e0e0'),
        top=Side(style='thin', color='e0e0e0'),
        bottom=Side(style='thin', color='e0e0e0')
    )
    for col in range(1, ws.max_column + 1):
        ws.cell(1, col).fill = header_fill
        ws.cell(1, col).font = header_font
        ws.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(1, col).border = thin_border
        ws.cell(2, col).fill = subheader_fill
        ws.cell(2, col).font = subheader_font
        ws.cell(2, col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(2, col).border = thin_border
    for row_idx in range(3, ws.max_row + 1):
        current_col = 1
        ws.cell(row_idx, current_col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row_idx, current_col).border = thin_border
        current_col += 1
        for grade in existing_grades:
            ws.cell(row_idx, current_col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row_idx, current_col).border = thin_border
            ws.cell(row_idx, current_col).font = Font(size=11, bold=True)
            ws.cell(row_idx, current_col + 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row_idx, current_col + 1).border = thin_border
            ws.cell(row_idx, current_col + 1).font = Font(color='6b6b6b', size=11)
            ws.cell(row_idx, current_col + 2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row_idx, current_col + 2).border = thin_border
            ws.cell(row_idx, current_col + 2).fill = yellow_fill
            ws.cell(row_idx, current_col + 2).font = Font(color='1565c0', size=11, bold=True)
            current_col += 3
    ws.column_dimensions['A'].width = 18
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
    
    # 시트 2: 승객
    ws2 = wb.create_sheet(title='승객')
    current_col = 1
    ws2.cell(1, current_col, '날짜')
    ws2.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
    current_col += 1
    ws2.cell(1, current_col, '출발')
    ws2.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
    current_col += 1
    ws2.cell(1, current_col, '도착')
    ws2.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
    current_col += 1
    for grade in existing_grades:
        ws2.cell(1, current_col, grade)
        ws2.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 2)
        current_col += 3
    current_col = 4
    for grade in existing_grades:
        ws2.cell(2, current_col, '확정')
        ws2.cell(2, current_col + 1, '블록')
        ws2.cell(2, current_col + 2, '잔여')
        current_col += 3
    for row_idx, row in passenger_final_df.iterrows():
        excel_row = row_idx + 3
        current_col = 1
        ws2.cell(excel_row, current_col, row['날짜'])
        current_col += 1
        ws2.cell(excel_row, current_col, row.get('departure_port', '-'))
        current_col += 1
        ws2.cell(excel_row, current_col, row.get('arrival_port', '-'))
        current_col += 1
        for grade in existing_grades:
            ws2.cell(excel_row, current_col, int(row.get(f'{grade}_확정', 0)))
            ws2.cell(excel_row, current_col + 1, int(row.get(f'{grade}_블록', 0)))
            ws2.cell(excel_row, current_col + 2, int(row.get(f'{grade}_잔여', 0)))
            current_col += 3
    for col in range(1, ws2.max_column + 1):
        ws2.cell(1, col).fill = header_fill
        ws2.cell(1, col).font = header_font
        ws2.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(1, col).border = thin_border
        ws2.cell(2, col).fill = subheader_fill
        ws2.cell(2, col).font = subheader_font
        ws2.cell(2, col).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(2, col).border = thin_border
    for row_idx in range(3, ws2.max_row + 1):
        current_col = 1
        ws2.cell(row_idx, current_col).alignment = Alignment(horizontal='left', vertical='center')
        ws2.cell(row_idx, current_col).border = thin_border
        current_col += 1
        for grade in existing_grades:
            ws2.cell(row_idx, current_col).alignment = Alignment(horizontal='center', vertical='center')
            ws2.cell(row_idx, current_col).border = thin_border
            ws2.cell(row_idx, current_col).font = Font(size=11, bold=True)
            ws2.cell(row_idx, current_col + 1).alignment = Alignment(horizontal='center', vertical='center')
            ws2.cell(row_idx, current_col + 1).border = thin_border
            ws2.cell(row_idx, current_col + 1).font = Font(color='6b6b6b', size=11)
            ws2.cell(row_idx, current_col + 2).alignment = Alignment(horizontal='center', vertical='center')
            ws2.cell(row_idx, current_col + 2).border = thin_border
            ws2.cell(row_idx, current_col + 2).fill = yellow_fill
            ws2.cell(row_idx, current_col + 2).font = Font(color='1565c0', size=11, bold=True)
            current_col += 3
    ws2.column_dimensions['A'].width = 18
    for col_idx in range(2, ws2.max_column + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 10
    ws2.row_dimensions[1].height = 25
    ws2.row_dimensions[2].height = 20
    for row_idx in range(3, ws2.max_row + 1):
        ws2.row_dimensions[row_idx].height = 20
    
    # 시트 3: 생성처별 (국적 기준)
    ws3 = wb.create_sheet(title='생성처별')
    df_passenger_analysis = result.get('passenger_analysis', pd.DataFrame())
    df_schedules_excel = result.get('schedules', pd.DataFrame())
    
    if not df_passenger_analysis.empty:
        # 국적 분류 함수
        def get_nationality_group_excel(nationality):
            if pd.isna(nationality) or not nationality:
                return '기타 국적'
            nat_upper = str(nationality).upper()
            if nat_upper == 'KR':
                return '한국 국적'
            elif nat_upper == 'JP':
                return '일본 국적'
            else:
                return '기타 국적'
        
        df_origin_excel = df_passenger_analysis.copy()
        df_origin_excel['nationality_group'] = df_origin_excel['nationality'].apply(get_nationality_group_excel)
        
        # 도착 포트 계산 - direction 기반으로 통일
        route_ports_info = route_direction_map.get(selected_route, {'first': '-', 'second': '-'})
        first_port_excel = route_ports_info.get('first', '-')
        second_port_excel = route_ports_info.get('second', '-')
        
        if not df_schedules_excel.empty and 'direction' in df_schedules_excel.columns:
            schedule_direction_map_excel = df_schedules_excel.set_index('schedule_id')['direction'].to_dict()
            df_origin_excel['direction'] = df_origin_excel['schedule_id'].map(schedule_direction_map_excel)
            
            df_origin_excel['arrival_port'] = df_origin_excel['direction'].apply(
                lambda d: second_port_excel if d == 'E' else (first_port_excel if d == 'W' else '-')
            )
        else:
            df_origin_excel['arrival_port'] = '-'
        
        # 스케줄+도착포트별 국적 집계
        if 'arrival_port' in df_origin_excel.columns:
            origin_summary_excel = df_origin_excel.groupby(['schedule_id', 'arrival_port', 'nationality_group']).size().unstack(fill_value=0).reset_index()
        else:
            origin_summary_excel = df_origin_excel.groupby(['schedule_id', 'nationality_group']).size().unstack(fill_value=0).reset_index()
            origin_summary_excel['arrival_port'] = '-'
        
        for col in ['한국 국적', '일본 국적', '기타 국적']:
            if col not in origin_summary_excel.columns:
                origin_summary_excel[col] = 0
        
        # 스케줄 정보 병합
        if not df_schedules_excel.empty:
            schedule_cols_excel = ['schedule_id', 'date', 'time_display', 'departure_port']
            available_cols_excel = [c for c in schedule_cols_excel if c in df_schedules_excel.columns]
            schedule_info_excel = df_schedules_excel[available_cols_excel].drop_duplicates()
            origin_summary_excel = origin_summary_excel.merge(schedule_info_excel, on='schedule_id', how='left')
            
            if 'date' in origin_summary_excel.columns:
                origin_summary_excel['date_display'] = pd.to_datetime(origin_summary_excel['date']).dt.strftime('%m-%d')
                weekday_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
                origin_summary_excel['weekday'] = pd.to_datetime(origin_summary_excel['date']).dt.dayofweek.map(weekday_map)
            
            origin_summary_excel['총계'] = origin_summary_excel['한국 국적'] + origin_summary_excel['일본 국적'] + origin_summary_excel['기타 국적']
            
            if 'date' in origin_summary_excel.columns and 'time_display' in origin_summary_excel.columns:
                origin_summary_excel = origin_summary_excel.sort_values(['date', 'time_display', 'arrival_port'])
        
        # 헤더
        ws3.cell(1, 1, '날짜')
        ws3.cell(1, 2, '출발')
        ws3.cell(1, 3, '도착')
        ws3.cell(1, 4, '한국 국적')
        ws3.cell(1, 5, '일본 국적')
        ws3.cell(1, 6, '기타 국적')
        ws3.cell(1, 7, '총계')
        
        for col in range(1, 8):
            ws3.cell(1, col).fill = header_fill
            ws3.cell(1, col).font = header_font
            ws3.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')
            ws3.cell(1, col).border = thin_border
        
        # 데이터
        for row_idx, row in origin_summary_excel.iterrows():
            excel_row = row_idx + 2
            time_str = row.get('time_display', '') or ''
            date_str = f"{row.get('date_display', '')} {time_str} ({row.get('weekday', '')})"
            dep_port = row.get('departure_port', '-') or '-'
            arr_port = row.get('arrival_port', '-') or '-'
            
            ws3.cell(excel_row, 1, date_str)
            ws3.cell(excel_row, 2, dep_port)
            ws3.cell(excel_row, 3, arr_port)
            ws3.cell(excel_row, 4, int(row.get('한국 국적', 0)))
            ws3.cell(excel_row, 5, int(row.get('일본 국적', 0)))
            ws3.cell(excel_row, 6, int(row.get('기타 국적', 0)))
            ws3.cell(excel_row, 7, int(row.get('총계', 0)))
            
            for col in range(1, 8):
                ws3.cell(excel_row, col).alignment = Alignment(horizontal='center', vertical='center')
                ws3.cell(excel_row, col).border = thin_border
            
            # 한국: 파란색, 일본: 빨간색
            ws3.cell(excel_row, 4).font = Font(color='436CFC', bold=True)
            ws3.cell(excel_row, 5).font = Font(color='EA3336', bold=True)
            ws3.cell(excel_row, 7).font = Font(bold=True)
        
        # 합계 row 추가
        total_row = len(origin_summary_excel) + 2
        ws3.cell(total_row, 1, '합계')
        ws3.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
        ws3.cell(total_row, 4, int(origin_summary_excel['한국 국적'].sum()))
        ws3.cell(total_row, 5, int(origin_summary_excel['일본 국적'].sum()))
        ws3.cell(total_row, 6, int(origin_summary_excel['기타 국적'].sum()))
        ws3.cell(total_row, 7, int(origin_summary_excel['총계'].sum()))
        
        for col in range(1, 8):
            ws3.cell(total_row, col).fill = header_fill
            ws3.cell(total_row, col).font = Font(color='FFFFFF', bold=True)
            ws3.cell(total_row, col).alignment = Alignment(horizontal='center', vertical='center')
            ws3.cell(total_row, col).border = thin_border
        
        ws3.column_dimensions['A'].width = 22
        ws3.column_dimensions['B'].width = 10
        ws3.column_dimensions['C'].width = 10
        ws3.column_dimensions['D'].width = 10
        ws3.column_dimensions['E'].width = 10
        ws3.column_dimensions['F'].width = 10
        ws3.column_dimensions['G'].width = 10
        ws3.row_dimensions[1].height = 25
    
    output = io.BytesIO()
    wb.save(output)
    excel_data = output.getvalue()
    
    # 탭 + 엑셀 버튼을 같은 줄에 배치 (CSS로 조정)
    st.markdown("""
    <style>
    .tab-header-container {
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    div[data-testid="stTabs"] {
        position: relative;
    }
    .excel-btn-wrapper {
        position: absolute;
        right: 0;
        top: 0;
        z-index: 100;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # 엑셀 버튼을 가장 오른쪽에 배치
    col_spacer, col_excel = st.columns([10, 1])
    with col_excel:
        st.download_button(
            label="엑셀 출력",
            data=excel_data,
            file_name=f"크루즈현황_{start_date}_{end_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="excel_download_top"
        )
    
    # 탭 상태 초기화
    if 'selected_tab' not in st.session_state:
        st.session_state.selected_tab = tab1_name
    
    # 탭 옵션
    tab_options = [tab1_name, "승객", "📊 승객 분석", "📍 생성처별 분석"]
    
    # 현재 선택된 탭이 옵션에 없으면 기본값으로
    if st.session_state.selected_tab not in tab_options:
        st.session_state.selected_tab = tab1_name
    
    # 선택된 탭 인덱스
    selected_idx = tab_options.index(st.session_state.selected_tab)
    
    # 진짜 st.tabs 표시
    tab_placeholder = st.tabs(tab_options)
    
    # JavaScript로 저장된 탭으로 클릭 + 탭 클릭 감지
    st.components.v1.html(f"""
        <script>
            setTimeout(function() {{
                // 저장된 탭으로 이동
                const tabs = window.parent.document.querySelectorAll('[data-baseweb="tab"]');
                if (tabs && tabs.length > {selected_idx}) {{
                    tabs[{selected_idx}].click();
                }}
                
                // 탭 클릭 시 숨겨진 버튼 클릭
                tabs.forEach((tab, idx) => {{
                    tab.addEventListener('click', function(e) {{
                        const buttons = window.parent.document.querySelectorAll('button[kind="secondary"]');
                        // 숨겨진 버튼 찾아서 클릭
                        buttons.forEach(btn => {{
                            if (btn.innerText === ['{ tab_options[0] }', '{ tab_options[1] }', '{ tab_options[2] }', '{ tab_options[3] }'][idx]) {{
                                btn.click();
                            }}
                        }});
                    }});
                }});
            }}, 50);
        </script>
    """, height=0)
    
    # 숨겨진 버튼들 (탭 전환용) - CSS로 화면 밖으로 이동
    st.markdown("""
        <style>
        /* 숨겨진 탭 버튼들 - 여러 선택자로 확실히 숨김 */
        div[data-testid="column"] button[kind="secondary"],
        button[kind="secondary"][data-testid="stBaseButton-secondary"],
        .stButton > button[kind="secondary"],
        div.row-widget button[kind="secondary"] {
            position: absolute !important;
            left: -9999px !important;
            width: 1px !important;
            height: 1px !important;
            overflow: hidden !important;
        }
        </style>
    """, unsafe_allow_html=True)
    btn_cols = st.columns(4)
    for idx, (col, tab_name) in enumerate(zip(btn_cols, tab_options)):
        with col:
            if st.button(tab_name, key=f"hidden_tab_btn_{idx}"):
                st.session_state.selected_tab = tab_name
                st.rerun()
    
    # 현재 선택된 탭
    selected_tab = st.session_state.selected_tab
    
    if selected_tab == tab1_name:
        # 객실 테이블 렌더링 (JavaScript 모달 포함)
        room_data_json = json.dumps(result.get('room_details', []), ensure_ascii=False)
        
        # 테이블 + 모달 + JavaScript를 하나의 HTML로 합침 (NEOHELIOS 디자인)
        full_html = f'''
        <meta charset="UTF-8">
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap" rel="stylesheet">
        <style>
            * {{
                font-family: 'Noto Sans KR', -apple-system, BlinkMacSystemFont, sans-serif;
                letter-spacing: -0.5px;
            }}
            
            /* 모달 오버레이 */
            #js-modal-overlay {{
                display: none;
                position: fixed;
                top: 0;
                left: 0;
                width: 100vw;
                height: 100vh;
                background: rgba(14, 14, 44, 0.5);
                z-index: 999998;
            }}
            #js-modal-overlay.show {{
                display: flex;
                justify-content: center;
                align-items: flex-start;
                padding-top: 20px;
            }}
            
            /* 모달 박스 */
            #js-modal-box {{
                background: #FFFFFF;
                border-radius: 5px;
                width: 95%;
                max-width: 800px;
                max-height: 90vh;
                overflow: hidden;
                box-shadow: 0 4px 24px rgba(14, 14, 44, 0.15);
            }}
            
            /* 모달 헤더 */
            #js-modal-header {{
                background: #232A5E;
                color: #FAFCFE;
                padding: 16px 20px;
                display: flex;
                justify-content: space-between;
                align-items: center;
            }}
            #js-modal-header h3 {{
                margin: 0;
                font-size: 16px;
                font-weight: 700;
            }}
            #js-modal-close {{
                background: none;
                border: none;
                color: #FAFCFE;
                font-size: 24px;
                cursor: pointer;
                padding: 4px 8px;
                line-height: 1;
                border-radius: 4px;
                transition: background 0.2s;
            }}
            #js-modal-close:hover {{
                background: rgba(255, 255, 255, 0.2);
            }}
            
            /* 모달 바디 */
            #js-modal-body {{
                padding: 20px;
                max-height: 70vh;
                overflow-y: auto;
            }}
            #js-modal-body table {{
                width: 100%;
                border-collapse: collapse;
            }}
            #js-modal-body th {{
                background: #232A5E;
                color: #FAFCFE;
                padding: 10px 12px;
                border: none;
                text-align: center;
                font-weight: 700;
                font-size: 12px;
            }}
            #js-modal-body td {{
                padding: 10px 12px;
                border-bottom: 1px solid #DAE0E3;
                text-align: center;
                font-size: 14px;
                color: #0E0E2C;
            }}
            #js-modal-body tr:nth-child(even) {{
                background: #F9FAFB;
            }}
            #js-modal-body tr:hover {{
                background: #F3F7F9;
            }}
            
            /* 클릭 가능 셀 */
            .clickable-cell:hover {{
                background: #F3F6FF !important;
                cursor: pointer;
            }}
            
            /* 모바일 가로 스크롤 */
            .responsive-table-container {{
                width: 100%;
                overflow-x: auto;
                overflow-y: auto;
                max-height: calc(100vh - 100px);
                -webkit-overflow-scrolling: touch;
                border-radius: 5px;
                border: 1px solid #DAE0E3;
            }}
            
            /* 헤더 고정 (상하 스크롤 시) */
            .responsive-table-container thead th {{
                position: sticky;
                top: 0;
                z-index: 10;
            }}
            .responsive-table-container thead tr:first-child th {{
                top: 0;
            }}
            .responsive-table-container thead tr:nth-child(2) th {{
                top: 46px;
            }}
            
            /* 첫 번째 열 고정 (좌우 스크롤 시) */
            .sticky-date-header {{
                position: sticky !important;
                left: 0 !important;
                z-index: 20 !important;
                background: #232A5E !important;
            }}
            .sticky-date-cell {{
                position: sticky !important;
                left: 0 !important;
                z-index: 5 !important;
            }}
        </style>
        
        <!-- 테이블 -->
        {result['html_table']}
        
        <!-- 모달 HTML -->
        <div id="js-modal-overlay" onclick="if(event.target.id==='js-modal-overlay') closeRoomModal()">
            <div id="js-modal-box">
                <div id="js-modal-header">
                    <h3 id="js-modal-title">객실 상세</h3>
                    <button id="js-modal-close" onclick="closeRoomModal()">&times;</button>
                </div>
                <div id="js-modal-body">
                    <p>로딩 중...</p>
                </div>
            </div>
        </div>
        
        <script>
            // 객실 상세 데이터
            const roomData = {room_data_json};
            
            // 디버깅용 로그 - status 분포 확인
            console.log('=== roomData 디버깅 ===');
            console.log('총 개수:', roomData.length);
            const statusCount = {{}};
            roomData.forEach(r => {{
                statusCount[r.status] = (statusCount[r.status] || 0) + 1;
            }});
            console.log('status 분포:', statusCount);
            console.log('샘플 (처음 10개):', roomData.slice(0, 10));
            
            // status 영어 -> 한글 변환
            const statusMap = {{
                'confirmed': '확정',
                'blocked': '블록',
                'vacant': '공실'
            }};
            
            // 모달 열기
            function openRoomModal(scheduleId, dateStr, grade, status) {{
                console.log('=== 클릭 이벤트 ===');
                console.log('scheduleId:', scheduleId, typeof scheduleId);
                console.log('grade:', grade);
                console.log('status:', status);
                
                // 해당 schedule_id의 모든 데이터 확인
                const sameSchedule = roomData.filter(r => String(r.schedule_id) === String(scheduleId));
                console.log('같은 schedule_id 데이터:', sameSchedule.length, '개');
                console.log('같은 schedule_id의 status들:', [...new Set(sameSchedule.map(r => r.status))]);
                
                // 타입 안전한 비교 (문자열로 변환하여 비교)
                const filtered = roomData.filter(r => 
                    String(r.schedule_id) === String(scheduleId) && 
                    r.grade === grade && 
                    r.status === status
                );
                
                console.log('최종 필터 결과:', filtered.length, '개');
                
                // 제목에는 한글로 표시
                const statusKo = statusMap[status] || status;
                document.getElementById('js-modal-title').textContent = dateStr + ' | ' + grade + ' | ' + statusKo;
                
                let html = '';
                if (filtered.length > 0) {{
                    html = '<table><tr><th>순번</th><th>객실등급</th><th>객실번호</th></tr>';
                    filtered.forEach((room, idx) => {{
                        html += '<tr><td>' + (idx + 1) + '</td><td>' + room.grade + '</td><td>' + room.room_no + '</td></tr>';
                    }});
                    html += '</table>';
                }} else {{
                    html = '<p style="text-align:center; color:#666; padding:30px;">해당 조건의 객실이 없습니다.</p>';
                }}
                
                document.getElementById('js-modal-body').innerHTML = html;
                document.getElementById('js-modal-overlay').classList.add('show');
            }}
            
            // 모달 닫기
            function closeRoomModal() {{
                document.getElementById('js-modal-overlay').classList.remove('show');
            }}
            
            // ESC 키로 닫기
            document.addEventListener('keydown', function(e) {{
                if (e.key === 'Escape') closeRoomModal();
            }});
        </script>
        '''
        
        # 테이블 행 수에 따라 높이 계산 (세로는 전체 표시, 가로는 스크롤 가능)
        row_count = len(result.get('final_df', []))
        # 헤더 2행 (약 120px) + 각 데이터 행 (약 65px) + 모달 여유 공간 (500px)
        table_height = 200 + row_count * 65 + 500
        
        # scrolling=True로 모바일 가로 스크롤 허용
        components.html(full_html, height=table_height, scrolling=True)
        
        # 범례 - 객실용 (NEOHELIOS 디자인)
        st.markdown("""
        <div style="margin-top: 24px; padding: 20px; background: #FFFFFF; border-radius: 5px; border: 1px solid #DAE0E3; font-family: 'Noto Sans KR', sans-serif;">
            <div style="color: #232A5E; font-weight: 700; font-size: 12px; margin-bottom: 16px; letter-spacing: -0.5px;">범례</div>
            <div class="legend-container" style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 16px;">
                <div style="display: flex; align-items: center;">
                    <span style="display: inline-block; width: 20px; height: 20px; background: #0E0E2C; border-radius: 3px; margin-right: 10px;"></span>
                    <span style="color: #0E0E2C; font-size: 14px; font-weight: 500;">확정 (명단 입력 완료)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="display: inline-block; width: 20px; height: 20px; background: #88949C; border-radius: 3px; margin-right: 10px;"></span>
                    <span style="color: #88949C; font-size: 14px; font-weight: 500;">블록 (점유 상태)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="display: inline-block; width: 20px; height: 20px; background: #FFFBEB; border: 1px solid #436CFC; border-radius: 3px; margin-right: 10px;"></span>
                    <span style="color: #436CFC; font-size: 14px; font-weight: 500;">공실 (예약 가능)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="display: inline-block; width: 20px; height: 20px; background: #EA3336; border-radius: 3px; margin-right: 10px;"></span>
                    <span style="color: #EA3336; font-size: 14px; font-weight: 700;">예약불가 (공실 0개)</span>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
                
    elif selected_tab == "승객":
        # 승객 테이블 생성
        final_df_passengers = result['final_df_passengers']
        existing_grades = result['existing_grades']
        
        # 승객 테이블 HTML 생성 (NEOHELIOS 디자인)
        html_pass_table = '<div class="responsive-table-container"><table style="width: 100%; border-collapse: collapse; background: #FFFFFF; font-family: Noto Sans KR, sans-serif;">'
        
        # 헤더 1행: 등급명
        html_pass_table += '<thead><tr><th rowspan="2" class="sticky-date-header" style="background: #232A5E; color: #FAFCFE; padding: 12px 10px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; letter-spacing: -0.5px; text-align: center;">날짜</th>'
        html_pass_table += '<th rowspan="2" style="background: #232A5E; color: #FAFCFE; padding: 12px 8px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; letter-spacing: -0.5px; text-align: center;">출발</th>'
        html_pass_table += '<th rowspan="2" style="background: #232A5E; color: #FAFCFE; padding: 12px 8px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; letter-spacing: -0.5px; text-align: center;">도착</th>'
        for idx, grade in enumerate(existing_grades):
            if grade == '총계':
                bg_color = '#1a2148'
            else:
                bg_color = '#232A5E'
            
            is_last_grade = (idx == len(existing_grades) - 1)
            border_right = '1px solid #3a4a7e' if is_last_grade else '1px solid #3a4a7e'
            
            html_pass_table += f'<th colspan="3" style="background: {bg_color}; color: #FAFCFE; padding: 12px 10px; border: none; border-right: {border_right}; font-weight: 700; font-size: 12px; letter-spacing: -0.5px; text-align: center;">{grade}</th>'
        html_pass_table += '</tr>'
        
        # 헤더 2행: 확정/블록/잔여
        html_pass_table += '<tr>'
        for idx, grade in enumerate(existing_grades):
            is_last_grade = (idx == len(existing_grades) - 1)
            grade_separator = '1px solid #DAE0E3' if is_last_grade else '1px solid #c8d0d4'
            
            html_pass_table += '<th style="background: #F3F7F9; color: #232A5E; text-align: center; padding: 10px 8px; font-weight: 500; border: none; border-right: 1px solid #DAE0E3; border-bottom: 1px solid #DAE0E3; font-size: 12px; letter-spacing: -0.5px;">확정</th>'
            html_pass_table += '<th style="background: #F3F7F9; color: #232A5E; text-align: center; padding: 10px 8px; font-weight: 500; border: none; border-right: 1px solid #DAE0E3; border-bottom: 1px solid #DAE0E3; font-size: 12px; letter-spacing: -0.5px;">블록</th>'
            html_pass_table += f'<th style="background: #FFFBEB; color: #232A5E; text-align: center; padding: 10px 8px; font-weight: 500; border: none; border-right: {grade_separator}; border-bottom: 1px solid #DAE0E3; font-size: 12px; letter-spacing: -0.5px;">잔여</th>'
        html_pass_table += '</tr></thead>'
        
        # 바디
        html_pass_table += '<tbody>'
        for idx, row in final_df_passengers.iterrows():
            row_bg = '#FFFFFF' if idx % 2 == 0 else '#F9FAFB'
            dep_port = row.get('departure_port', '-')
            arr_port = row.get('arrival_port', '-')
            
            html_pass_table += '<tr style="border-bottom: 1px solid #DAE0E3; transition: background 0.15s ease;">'
            html_pass_table += f'<td class="sticky-date-cell" style="background: {row_bg}; color: #0E0E2C; font-weight: 500; padding: 10px; border: none; border-right: 1px solid #DAE0E3; font-size: 14px; letter-spacing: -0.5px; text-align: center;">{row["날짜"]}</td>'
            html_pass_table += f'<td style="background: {row_bg}; color: #0E0E2C; font-weight: 500; padding: 10px 8px; border: none; border-right: 1px solid #DAE0E3; font-size: 14px; letter-spacing: -0.5px; text-align: center;">{dep_port}</td>'
            html_pass_table += f'<td style="background: {row_bg}; color: #0E0E2C; font-weight: 500; padding: 10px 8px; border: none; border-right: 1px solid #DAE0E3; font-size: 14px; letter-spacing: -0.5px; text-align: center;">{arr_port}</td>'
            
            for idx_g, grade in enumerate(existing_grades):
                confirmed = int(row.get(f'{grade}_확정', 0))
                blocked = int(row.get(f'{grade}_블록', 0))
                remaining = int(row.get(f'{grade}_잔여', 0))
                
                is_last_grade = (idx_g == len(existing_grades) - 1)
                grade_separator = '1px solid #DAE0E3' if is_last_grade else '1px solid #c8d0d4'
                
                # 확정
                html_pass_table += f'<td style="background: {row_bg}; color: #0E0E2C; text-align: center; padding: 10px; font-weight: 500; border: none; border-right: 1px solid #DAE0E3; font-size: 14px;">{confirmed}</td>'
                
                # 블록
                html_pass_table += f'<td style="background: {row_bg}; color: #88949C; text-align: center; padding: 10px; font-weight: 400; border: none; border-right: 1px solid #DAE0E3; font-size: 14px;">{blocked}</td>'
                
                # 잔여 (노란색 배경, 0이면 빨간색)
                if remaining == 0:
                    remaining_style = f'background: #FEF2F2; color: #EA3336; text-align: center; padding: 10px; font-weight: 700; border: none; border-right: {grade_separator}; border-left: 3px solid #EA3336; font-size: 14px;'
                else:
                    yellow_bg = '#FFFBEB' if row_bg == '#FFFFFF' else '#FEF9E7'
                    remaining_style = f'background: {yellow_bg}; color: #436CFC; text-align: center; padding: 10px; font-weight: 500; border: none; border-right: {grade_separator}; font-size: 14px;'
                html_pass_table += f'<td style="{remaining_style}">{remaining}</td>'
            
            html_pass_table += '</tr>'
        html_pass_table += '</tbody></table></div>'
        
        st.markdown(html_pass_table, unsafe_allow_html=True)
        
        # 범례 - 승객용 (NEOHELIOS 디자인)
        st.markdown("""
        <div style="margin-top: 24px; padding: 20px; background: #FFFFFF; border-radius: 5px; border: 1px solid #DAE0E3; font-family: 'Noto Sans KR', sans-serif;">
            <div style="color: #232A5E; font-weight: 700; font-size: 12px; margin-bottom: 16px; letter-spacing: -0.5px;">범례</div>
            <div class="legend-container" style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 16px;">
                <div style="display: flex; align-items: center;">
                    <span style="display: inline-block; width: 20px; height: 20px; background: #0E0E2C; border-radius: 3px; margin-right: 10px;"></span>
                    <span style="color: #0E0E2C; font-size: 14px; font-weight: 500;">확정 (명단 입력 완료)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="display: inline-block; width: 20px; height: 20px; background: #88949C; border-radius: 3px; margin-right: 10px;"></span>
                    <span style="color: #88949C; font-size: 14px; font-weight: 500;">블록 (점유 상태, 정원 제한 적용)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="display: inline-block; width: 20px; height: 20px; background: #FFFBEB; border: 1px solid #436CFC; border-radius: 3px; margin-right: 10px;"></span>
                    <span style="color: #436CFC; font-size: 14px; font-weight: 500;">잔여 (예약 가능 인원)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="display: inline-block; width: 20px; height: 20px; background: #EA3336; border-radius: 3px; margin-right: 10px;"></span>
                    <span style="color: #EA3336; font-size: 14px; font-weight: 700;">예약불가 (잔여 0명)</span>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    elif selected_tab == "📊 승객 분석":
        # 승객 분석 대시보드
        df_analysis = result.get('passenger_analysis', pd.DataFrame())
        
        if df_analysis.empty:
            st.info("확정된 승객 데이터가 없습니다.")
        else:
            # 승객 분석 탭에 진입했음을 기록 (탭 상태 유지용)
            st.session_state.selected_tab = 2
            
            # 승객 분석 전용 필터
            filter_col1, filter_col2, _ = st.columns([2, 2, 6])
            with filter_col1:
                issue_options = ['전체', '발권완료', '미발권']
                selected_issue_status = st.selectbox("발권 상태", issue_options, index=0, key="issue_status_select")
            with filter_col2:
                origin_country_options = ['전체', '한국', '일본']
                selected_origin_country = st.selectbox("생성처", origin_country_options, index=0, key="origin_country_select")
            
            # 발권 상태로 필터링
            if selected_issue_status == '발권완료':
                df_analysis = df_analysis[df_analysis['is_issued'] == 1].copy()
            elif selected_issue_status == '미발권':
                df_analysis = df_analysis[df_analysis['is_issued'] == 0].copy()
            
            # 생성처로 필터링 (ticket_number 첫 글자: K=한국, J=일본)
            if 'ticket_number' in df_analysis.columns:
                if selected_origin_country == '한국':
                    df_analysis = df_analysis[df_analysis['ticket_number'].str.startswith('K', na=False)].copy()
                elif selected_origin_country == '일본':
                    df_analysis = df_analysis[df_analysis['ticket_number'].str.startswith('J', na=False)].copy()
            elif selected_origin_country != '전체':
                st.warning("생성처 필터를 적용하려면 다시 조회해주세요.")
            
            if df_analysis.empty:
                st.info(f"선택한 조건에 해당하는 승객 데이터가 없습니다.")
                st.stop()
            
            # 데이터 전처리
            today = datetime.today()
            
            # 연령 계산 (birth_day가 있는 경우만)
            df_analysis = df_analysis.dropna(subset=['birth_day'])
            df_analysis['birth_day'] = pd.to_datetime(df_analysis['birth_day'])
            df_analysis['age'] = df_analysis['birth_day'].apply(
                lambda x: (today - x).days // 365 if pd.notna(x) else None
            )
            
            # 연령대 분류
            def get_age_group(age):
                if age is None or pd.isna(age):
                    return '미상'
                elif age < 10:
                    return '0-9세'
                elif age < 20:
                    return '10대'
                elif age < 30:
                    return '20대'
                elif age < 40:
                    return '30대'
                elif age < 50:
                    return '40대'
                elif age < 60:
                    return '50대'
                elif age < 70:
                    return '60대'
                else:
                    return '70대+'
            
            df_analysis['age_group'] = df_analysis['age'].apply(get_age_group)
            
            # 국적 코드 -> 국가명 변환
            nationality_map = {
                'KR': '한국 🇰🇷',
                'JP': '일본 🇯🇵',
                'CN': '중국 🇨🇳',
                'US': '미국 🇺🇸',
                'TW': '대만 🇹🇼',
                'HK': '홍콩 🇭🇰',
                'VN': '베트남 🇻🇳',
                'TH': '태국 🇹🇭',
                'PH': '필리핀 🇵🇭',
                'MY': '말레이시아 🇲🇾',
                'SG': '싱가포르 🇸🇬',
                'ID': '인도네시아 🇮🇩',
                'AU': '호주 🇦🇺',
                'CA': '캐나다 🇨🇦',
                'GB': '영국 🇬🇧',
                'DE': '독일 🇩🇪',
                'FR': '프랑스 🇫🇷',
                'RU': '러시아 🇷🇺'
            }
            df_analysis['nationality_name'] = df_analysis['nationality'].map(
                lambda x: nationality_map.get(x, f'기타 ({x})') if pd.notna(x) else '미상'
            )
            
            # 성별 한글 변환
            sex_map = {'M': '남성', 'F': '여성'}
            df_analysis['sex_name'] = df_analysis['sex'].map(
                lambda x: sex_map.get(x, '미상') if pd.notna(x) else '미상'
            )
            
            # 총 승객 수
            total_passengers = len(df_analysis)
            
            # 헤더 (NEOHELIOS 디자인)
            st.markdown(f"""
            <div style="background: #232A5E; padding: 24px; border-radius: 5px; margin-bottom: 24px; font-family: 'Noto Sans KR', sans-serif;">
                <h2 style="color: #FAFCFE; margin: 0; font-size: 20px; font-weight: 700; letter-spacing: -0.5px;">
                    📊 승객 분석
                </h2>
                <p style="color: #9EA8B0; margin: 8px 0 0 0; font-size: 14px; letter-spacing: -0.5px;">
                    조회 기간: {result.get('start_date', '')} ~ {result.get('end_date', '')} | 
                    총 <span style="color: #436CFC; font-weight: 700;">{total_passengers:,}</span>명
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            # 3개 차트를 나란히 배치
            col1, col2, col3 = st.columns(3)
            
            # === 성별 분포 (도넛 차트) ===
            with col1:
                sex_counts = df_analysis['sex_name'].value_counts()
                
                # 성별에 따라 색상 매핑 (NEOHELIOS 디자인: 남성=파란색, 여성=분홍색)
                sex_colors = [('#F48FB1' if s == '여성' else '#436CFC' if s == '남성' else '#9EA8B0') for s in sex_counts.index]
                
                fig_sex = go.Figure(data=[go.Pie(
                    labels=sex_counts.index,
                    values=sex_counts.values,
                    hole=0.6,
                    marker=dict(
                        colors=sex_colors,
                        line=dict(color='#FFFFFF', width=2)
                    ),
                    textinfo='label+percent',
                    textfont=dict(size=12, family='Noto Sans KR'),
                    hovertemplate='%{label}<br>%{value}명 (%{percent})<extra></extra>'
                )])
                
                fig_sex.update_layout(
                    title=dict(
                        text='👤 성별 분포',
                        font=dict(size=16, color='#232A5E', family='Noto Sans KR'),
                        x=0.5
                    ),
                    showlegend=True,
                    legend=dict(
                        orientation='h',
                        yanchor='bottom',
                        y=-0.15,
                        xanchor='center',
                        x=0.5,
                        font=dict(family='Noto Sans KR', size=12, color='#0E0E2C')
                    ),
                    height=400,
                    margin=dict(t=60, b=60, l=20, r=20),
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='rgba(0,0,0,0)',
                    annotations=[dict(
                        text=f'<b>{total_passengers:,}</b><br>명',
                        x=0.5, y=0.5,
                        font=dict(size=18, color='#232A5E', family='Noto Sans KR'),
                        showarrow=False
                    )]
                )
                
                st.plotly_chart(fig_sex, use_container_width=True)
            
            # === 국적 분포 (가로 막대 차트) ===
            with col2:
                nationality_counts = df_analysis['nationality_name'].value_counts().head(10)
                
                # NEOHELIOS 색상 그라데이션 (많을수록 짙게)
                n_colors = len(nationality_counts)
                # 많은 순서대로 짙은 색 (1.0 -> 0.3)
                colors = [f'rgba(67, 108, 252, {1.0 - 0.7 * (i / max(n_colors-1, 1))})' for i in range(n_colors)]
                
                fig_nat = go.Figure(data=[go.Bar(
                    y=nationality_counts.index[::-1],
                    x=nationality_counts.values[::-1],
                    orientation='h',
                    marker=dict(
                        color=colors[::-1],
                        line=dict(color='#FFFFFF', width=1)
                    ),
                    text=nationality_counts.values[::-1],
                    textposition='outside',
                    textfont=dict(size=12, color='#0E0E2C', family='Noto Sans KR'),
                    hovertemplate='%{y}<br>%{x}명<extra></extra>'
                )])
                
                fig_nat.update_layout(
                    title=dict(
                        text='🌍 국적 분포 (Top 10)',
                        font=dict(size=16, color='#232A5E', family='Noto Sans KR'),
                        x=0.5
                    ),
                    xaxis=dict(
                        title=dict(text='승객 수', font=dict(family='Noto Sans KR', size=12, color='#88949C')),
                        showgrid=True,
                        gridcolor='#DAE0E3'
                    ),
                    yaxis=dict(
                        title='',
                        tickfont=dict(size=12)
                    ),
                    height=400,
                    margin=dict(t=60, b=40, l=120, r=40),
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='rgba(0,0,0,0)'
                )
                
                st.plotly_chart(fig_nat, use_container_width=True)
            
            # === 연령대 분포 (세로 막대 차트) ===
            with col3:
                # 연령대 순서 지정
                age_order = ['0-9세', '10대', '20대', '30대', '40대', '50대', '60대', '70대+', '미상']
                age_counts = df_analysis['age_group'].value_counts()
                age_counts = age_counts.reindex([a for a in age_order if a in age_counts.index])
                
                # 연령대별 색상 (젊은층: 밝은 색, 고령층: 진한 색)
                age_colors = ['#81d4fa', '#4fc3f7', '#29b6f6', '#03a9f4', '#039be5', '#0288d1', '#0277bd', '#01579b', '#b0bec5']
                
                fig_age = go.Figure(data=[go.Bar(
                    x=age_counts.index,
                    y=age_counts.values,
                    marker=dict(
                        color=age_colors[:len(age_counts)],
                        line=dict(color='#ffffff', width=1)
                    ),
                    text=age_counts.values,
                    textposition='outside',
                    textfont=dict(size=12, color='#333333'),
                    hovertemplate='%{x}<br>%{y}명<extra></extra>'
                )])
                
                fig_age.update_layout(
                    title=dict(
                        text='📈 연령대 분포',
                        font=dict(size=20, color='#333333'),
                        x=0.5
                    ),
                    xaxis=dict(
                        title='연령대',
                        tickfont=dict(size=11)
                    ),
                    yaxis=dict(
                        title='승객 수',
                        showgrid=True,
                        gridcolor='rgba(0,0,0,0.1)'
                    ),
                    height=400,
                    margin=dict(t=60, b=40, l=40, r=20),
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='rgba(0,0,0,0)'
                )
                
                st.plotly_chart(fig_age, use_container_width=True)
            
            # 상세 통계 테이블
            st.markdown("""
            <div style="margin-top: 20px; padding: 24px; background: #f8f9fa; border-radius: 8px;">
                <h4 style="color: #333; margin: 0 0 16px 0; font-size: 18px;">📋 상세 통계</h4>
            </div>
            """, unsafe_allow_html=True)
            
            stat_col1, stat_col2, stat_col3 = st.columns(3)
            
            with stat_col1:
                st.markdown("**성별 통계**")
                sex_df = df_analysis['sex_name'].value_counts().reset_index()
                sex_df.columns = ['성별', '인원']
                sex_df['비율'] = (sex_df['인원'] / sex_df['인원'].sum() * 100).round(1).astype(str) + '%'
                st.dataframe(sex_df, hide_index=True, use_container_width=True)
            
            with stat_col2:
                st.markdown("**국적 통계 (Top 10)**")
                nat_df = df_analysis['nationality_name'].value_counts().head(10).reset_index()
                nat_df.columns = ['국적', '인원']
                nat_df['비율'] = (nat_df['인원'] / total_passengers * 100).round(1).astype(str) + '%'
                st.dataframe(nat_df, hide_index=True, use_container_width=True)
            
            with stat_col3:
                st.markdown("**연령대 통계**")
                age_df = df_analysis['age_group'].value_counts().reindex([a for a in age_order if a in df_analysis['age_group'].value_counts().index]).reset_index()
                age_df.columns = ['연령대', '인원']
                age_df['비율'] = (age_df['인원'] / age_df['인원'].sum() * 100).round(1).astype(str) + '%'
                st.dataframe(age_df, hide_index=True, use_container_width=True)

    elif selected_tab == "📍 생성처별 분석":
        # 생성처별 분석 탭
        df_analysis = result.get('passenger_analysis', pd.DataFrame())
        df_schedules = result.get('schedules', pd.DataFrame())
        
        if df_analysis.empty:
            st.info("확정된 승객 데이터가 없습니다.")
        else:
            # 헤더
            st.markdown("""
            <div style="background: #232A5E; padding: 24px; border-radius: 5px; margin-bottom: 24px; font-family: 'Noto Sans KR', sans-serif;">
                <h2 style="color: #FAFCFE; margin: 0; font-size: 20px; font-weight: 700; letter-spacing: -0.5px;">
                    📍 생성처별 분석
                </h2>
                <p style="color: #9EA8B0; margin: 8px 0 0 0; font-size: 14px; letter-spacing: -0.5px;">
                    스케줄별 생성처(한국/일본) 승객 현황
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            # 생성처 필터 (K: 한국생성, J: 일본생성)
            origin_filter_col, _ = st.columns([2, 8])
            with origin_filter_col:
                origin_filter_options = ["전체", "한국", "일본"]
                selected_origin_filter = st.selectbox("생성처", origin_filter_options, index=0, key="origin_filter_tab4")
            
            # 생성처 분류 함수
            def get_origin(ticket_number):
                if pd.isna(ticket_number) or not ticket_number:
                    return '기타'
                first_char = str(ticket_number)[0].upper()
                if first_char == 'K':
                    return '한국'
                elif first_char == 'J':
                    return '일본'
                else:
                    return '기타'
            
            # 생성처 컬럼 추가
            df_origin = df_analysis.copy()
            if 'ticket_number' in df_origin.columns:
                df_origin['origin'] = df_origin['ticket_number'].apply(get_origin)
            else:
                df_origin['origin'] = '기타'
            
            # 생성처 필터 적용 (ticket_number 기반)
            if selected_origin_filter != "전체":
                df_origin = df_origin[df_origin['origin'] == selected_origin_filter].copy()
            
            # 국적 분류 (nationality 기반)
            def get_nationality_group(nationality):
                if pd.isna(nationality) or not nationality:
                    return '기타 국적'
                nat_upper = str(nationality).upper()
                if nat_upper == 'KR':
                    return '한국 국적'
                elif nat_upper == 'JP':
                    return '일본 국적'
                else:
                    return '기타 국적'
            
            df_origin['nationality_group'] = df_origin['nationality'].apply(get_nationality_group)
            
            # 도착 포트 계산 - direction 기반으로 통일 (더 안정적)
            route_ports_info = route_direction_map.get(selected_route, {'first': '-', 'second': '-'})
            first_port = route_ports_info.get('first', '-')
            second_port = route_ports_info.get('second', '-')
            
            if not df_schedules.empty and 'direction' in df_schedules.columns:
                schedule_direction_map = df_schedules.set_index('schedule_id')['direction'].to_dict()
                df_origin['direction'] = df_origin['schedule_id'].map(schedule_direction_map)
                
                # E방향: 첫번째→두번째, W방향: 두번째→첫번째
                df_origin['arrival_port'] = df_origin['direction'].apply(
                    lambda d: second_port if d == 'E' else (first_port if d == 'W' else '-')
                )
            else:
                df_origin['arrival_port'] = '-'
            
            # 스케줄+도착포트별 국적 집계
            if 'arrival_port' in df_origin.columns:
                origin_summary = df_origin.groupby(['schedule_id', 'arrival_port', 'nationality_group']).size().unstack(fill_value=0).reset_index()
            else:
                origin_summary = df_origin.groupby(['schedule_id', 'nationality_group']).size().unstack(fill_value=0).reset_index()
                origin_summary['arrival_port'] = '-'
            
            # 컬럼 정리
            for col in ['한국 국적', '일본 국적', '기타 국적']:
                if col not in origin_summary.columns:
                    origin_summary[col] = 0
            
            # 스케줄 정보 병합
            if not df_schedules.empty:
                schedule_cols = ['schedule_id', 'date', 'time_display', 'departure_port']
                available_cols = [c for c in schedule_cols if c in df_schedules.columns]
                schedule_info = df_schedules[available_cols].drop_duplicates()
                origin_summary = origin_summary.merge(schedule_info, on='schedule_id', how='left')
                
                # 날짜 포맷
                if 'date' in origin_summary.columns:
                    origin_summary['date_display'] = pd.to_datetime(origin_summary['date']).dt.strftime('%m-%d')
                    weekday_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
                    origin_summary['weekday'] = pd.to_datetime(origin_summary['date']).dt.dayofweek.map(weekday_map)
                else:
                    origin_summary['date_display'] = '-'
                    origin_summary['weekday'] = '-'
                
                # 총계 계산
                origin_summary['총계'] = origin_summary['한국 국적'] + origin_summary['일본 국적'] + origin_summary['기타 국적']
                
                # 정렬
                if 'date' in origin_summary.columns and 'time_display' in origin_summary.columns:
                    origin_summary = origin_summary.sort_values(['date', 'time_display', 'arrival_port'])
                
                # 테이블 HTML 생성
                html_origin = '<div class="responsive-table-container"><table style="width: 100%; border-collapse: collapse; background: #FFFFFF; font-family: Noto Sans KR, sans-serif;">'
                
                # 헤더
                html_origin += '''<thead><tr>
                    <th style="background: #232A5E; color: #FAFCFE; padding: 12px 10px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; text-align: center;">날짜</th>
                    <th style="background: #232A5E; color: #FAFCFE; padding: 12px 10px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; text-align: center;">출발</th>
                    <th style="background: #232A5E; color: #FAFCFE; padding: 12px 10px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; text-align: center;">도착</th>
                    <th style="background: #232A5E; color: #FAFCFE; padding: 12px 10px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; text-align: center;">한국 국적</th>
                    <th style="background: #232A5E; color: #FAFCFE; padding: 12px 10px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; text-align: center;">일본 국적</th>
                    <th style="background: #232A5E; color: #FAFCFE; padding: 12px 10px; border: none; border-right: 1px solid #3a4a7e; font-weight: 700; font-size: 12px; text-align: center;">기타 국적</th>
                    <th style="background: #1a2148; color: #FAFCFE; padding: 12px 10px; border: none; font-weight: 700; font-size: 12px; text-align: center;">총계</th>
                </tr></thead>'''
                
                # 바디
                html_origin += '<tbody>'
                row_idx = 0
                for _, row in origin_summary.iterrows():
                    row_bg = '#FFFFFF' if row_idx % 2 == 0 else '#F9FAFB'
                    time_str = row.get('time_display', '') or ''
                    date_str = f"{row.get('date_display', '')} {time_str} ({row.get('weekday', '')})"
                    dep_port = row.get('departure_port', '-') or '-'
                    arr_port = row.get('arrival_port', '-') or '-'
                    kr_count = int(row.get('한국 국적', 0))
                    jp_count = int(row.get('일본 국적', 0))
                    etc_count = int(row.get('기타 국적', 0))
                    total_count = int(row.get('총계', 0))
                    
                    html_origin += f'''<tr style="border-bottom: 1px solid #DAE0E3;">
                        <td style="background: {row_bg}; color: #0E0E2C; padding: 10px; border-right: 1px solid #DAE0E3; font-size: 14px; text-align: center;">{date_str}</td>
                        <td style="background: {row_bg}; color: #0E0E2C; padding: 10px; border-right: 1px solid #DAE0E3; font-size: 14px; text-align: center; font-weight: 500;">{dep_port}</td>
                        <td style="background: {row_bg}; color: #0E0E2C; padding: 10px; border-right: 1px solid #DAE0E3; font-size: 14px; text-align: center; font-weight: 500;">{arr_port}</td>
                        <td style="background: {row_bg}; color: #436CFC; padding: 10px; border-right: 1px solid #DAE0E3; font-size: 14px; text-align: center; font-weight: 600;">{kr_count}</td>
                        <td style="background: {row_bg}; color: #EA3336; padding: 10px; border-right: 1px solid #DAE0E3; font-size: 14px; text-align: center; font-weight: 600;">{jp_count}</td>
                        <td style="background: {row_bg}; color: #88949C; padding: 10px; border-right: 1px solid #DAE0E3; font-size: 14px; text-align: center;">{etc_count}</td>
                        <td style="background: {row_bg}; color: #232A5E; padding: 10px; font-size: 14px; text-align: center; font-weight: 700;">{total_count}</td>
                    </tr>'''
                    row_idx += 1
                
                # 합계 row 추가
                total_kr = origin_summary['한국 국적'].sum()
                total_jp = origin_summary['일본 국적'].sum()
                total_etc = origin_summary['기타 국적'].sum()
                total_all = total_kr + total_jp + total_etc
                
                html_origin += f'''<tr style="border-top: 2px solid #232A5E; background: #F3F6FF;">
                    <td colspan="3" style="background: #232A5E; color: #FAFCFE; padding: 12px 10px; border-right: 1px solid #3a4a7e; font-size: 14px; text-align: center; font-weight: 700;">합계</td>
                    <td style="background: #F3F6FF; color: #436CFC; padding: 12px 10px; border-right: 1px solid #DAE0E3; font-size: 14px; text-align: center; font-weight: 700;">{total_kr}</td>
                    <td style="background: #F3F6FF; color: #EA3336; padding: 12px 10px; border-right: 1px solid #DAE0E3; font-size: 14px; text-align: center; font-weight: 700;">{total_jp}</td>
                    <td style="background: #F3F6FF; color: #88949C; padding: 12px 10px; border-right: 1px solid #DAE0E3; font-size: 14px; text-align: center; font-weight: 600;">{total_etc}</td>
                    <td style="background: #232A5E; color: #FAFCFE; padding: 12px 10px; font-size: 14px; text-align: center; font-weight: 700;">{total_all}</td>
                </tr>'''
                
                html_origin += '</tbody></table></div>'
                
                st.markdown(html_origin, unsafe_allow_html=True)
            else:
                st.warning("스케줄 정보를 찾을 수 없습니다.")

st.markdown('<hr style="border: none; height: 1px; background: #DAE0E3; margin: 40px 0;">', unsafe_allow_html=True)
st.markdown('<p style="text-align: center; color: #999999; font-size: 12px;">문제가 있으면 DB 접속 정보를 확인하세요</p>', unsafe_allow_html=True)

